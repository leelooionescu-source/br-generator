# -*- coding: utf-8 -*-
"""
Procesare documentatii cadastrale - logica de extractie date din PDF-uri
si inserare in situatii Excel HG.

Suporta doua tipuri:
1. Scan complet PDF (un singur PDF per pozitie, ex: "POZITIA 1234.pdf")
2. Foldere cu EXPROPRIAT / RAMAS (ex: "1234_55001/EXPROPRIAT/PAD.pdf")
"""

import os
import re
import shutil
import zipfile
import traceback
from datetime import datetime
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# OCR support for scanned PDFs
try:
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image
    import io
    # Set Tesseract path - Windows vs Linux
    import platform
    if platform.system() == 'Windows':
        tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        if os.path.exists(tesseract_path):
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
    # On Linux (Render/Docker), tesseract is in PATH by default
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF, using OCR fallback for scanned pages."""
    full_text = ""
    try:
        pdf = pdfplumber.open(pdf_path)
        for page in pdf.pages:
            t = page.extract_text()
            if t and t.strip():
                full_text += t + "\n"
        pdf.close()
    except Exception:
        pass

    # If pdfplumber got text, return it
    if full_text.strip():
        return full_text

    # OCR fallback for scanned PDFs
    if not OCR_AVAILABLE:
        return full_text

    try:
        pdf = fitz.open(pdf_path)
        for page in pdf:
            mat = fitz.Matrix(300/72, 300/72)
            pix = page.get_pixmap(matrix=mat)
            img = Image.open(io.BytesIO(pix.tobytes('png')))
            text = pytesseract.image_to_string(img, lang='ron')
            if text:
                full_text += text + "\n"
        pdf.close()
    except Exception:
        pass

    return full_text


# ── Styles ──────────────────────────────────────────────────────────────
FONT_DEFAULT = Font(name='Arial', size=8, bold=True)
FONT_OBS_OK = Font(name='Arial', size=8, bold=True, color='FF0070C0')
FONT_OBS_DIFF = Font(name='Arial', size=8, bold=True, color='FF000000')
FONT_SUP_EXP = Font(name='Arial', size=8, bold=True, color='FF002060')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FILL_GREEN = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
FILL_RED = PatternFill(start_color='FFFF5050', end_color='FFFF5050', fill_type='solid')


# ── Excel helpers ───────────────────────────────────────────────────────
def sv(ws, r, c):
    """Safe value read from worksheet cell."""
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def ss(ws, r, c, val, font=None):
    """Safe set value in worksheet cell."""
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        return
    cell.value = val
    cell.font = font or FONT_DEFAULT
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN


def find_row(ws, poz):
    """Find row by position number in column A."""
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if not isinstance(cell, MergedCell) and cell.value is not None:
            try:
                if int(float(str(cell.value))) == poz:
                    return row
            except (ValueError, TypeError):
                continue
    return None


def is_row_filled(ws, row):
    """Check if doc cad columns are already filled."""
    for c in [19, 24, 26, 30]:
        cell = ws.cell(row=row, column=c)
        if not isinstance(cell, MergedCell) and cell.value is not None:
            return True
    return False


def color_row(ws, row, fill, max_col=35):
    """Apply fill color to entire row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if not isinstance(cell, MergedCell):
            cell.fill = fill


# ── Normalizare ─────────────────────────────────────────────────────────
def normalize_cat(raw):
    """Normalize land use category."""
    if raw is None:
        return raw
    r = str(raw).strip().upper()
    for old, new in [('Ă', 'A'), ('Â', 'A'), ('Î', 'I'), ('Ș', 'S'), ('Ț', 'T'), ('Ş', 'S'), ('Ţ', 'T')]:
        r = r.replace(old, new)
    if 'ARABIL' in r:
        return 'Arabil'
    if any(x in r for x in ['FANEA', 'FANET', 'FANATA', 'FANETE']):
        return 'Faneata'
    if 'PASUNE' in r:
        return 'Pasune'
    if 'PADURE' in r or 'SILVIC' in r:
        return 'Padure'
    if 'NEPRODUCTIV' in r or r == 'N':
        return 'Neproductiv'
    if 'CURTI' in r or 'CONSTRUCTII' in r or 'CURȚI' in r:
        return 'Curti constructii'
    if 'ALTELE' in r:
        return 'Altele'
    return raw.strip()


def parse_suprafata(s):
    """Parse suprafata string, handling Romanian thousands separator (dot)."""
    if s is None:
        return None
    s = str(s).strip()
    s = s.replace('.', '').replace(',', '')
    try:
        return int(s)
    except ValueError:
        return None


def compare_values(hg_val, doc_val):
    """Compare HG value with doc cad value, normalized."""
    if hg_val is None or doc_val is None:
        return True
    h = normalize_cat(str(hg_val))
    d = normalize_cat(str(doc_val))
    return h == d


def compare_numeric(hg_val, doc_val):
    """Compare numeric values."""
    if hg_val is None or doc_val is None:
        return True
    try:
        return int(float(str(hg_val))) == int(float(str(doc_val)))
    except (ValueError, TypeError):
        return True


# ── PDF Extraction ──────────────────────────────────────────────────────
def extract_cf_data(pdf_path):
    """Extract data from CF (carte funciara) extract PDF."""
    data = {
        'nr_cadastral': None,
        'nr_cf': None,
        'suprafata_totala': None,
        'categorie': None,
        'intravilan': None,
        'tarla': None,
        'parcela': None,
        'titular': None,
        'sarcini': [],
    }

    full_text = extract_text_from_pdf(pdf_path)
    if not full_text.strip():
        return data

    # Nr. CF
    m = re.search(r'Carte\s+Funciar[aă]\s+Nr\.?\s*(\d+)', full_text, re.IGNORECASE)
    if m:
        data['nr_cf'] = int(m.group(1))
        data['nr_cadastral'] = int(m.group(1))

    # Nr. cadastral from A1 row
    m = re.search(r'A1\s+(\d+)\s+', full_text)
    if m:
        data['nr_cadastral'] = int(m.group(1))

    # Suprafata totala - from A1 row
    m = re.search(r'A1\s+\d+\s+([\d.,]+)', full_text)
    if m:
        data['suprafata_totala'] = parse_suprafata(m.group(1))

    # Categorie de folosinta from "Date referitoare la teren" table
    teren_match = re.search(
        r'Date\s+referitoare\s+la\s+teren.*?'
        r'(\d+)\s+'
        r'([a-zA-ZăâîșțĂÂÎȘȚşţ\-\s]+?)\s+'
        r'(DA|NU)\s+'
        r'([\d.,]+)\s+'
        r'(\d+|[-–])\s+'
        r'([\d/]+|[-–])',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if teren_match:
        cat_raw = teren_match.group(2).strip()
        data['categorie'] = normalize_cat(cat_raw)
        data['intravilan'] = 'Intravilan' if teren_match.group(3).upper() == 'DA' else 'Extravilan'
        tarla = teren_match.group(5).strip()
        if tarla and tarla not in ['-', '–']:
            data['tarla'] = tarla
        parcela = teren_match.group(6).strip()
        if parcela and parcela not in ['-', '–']:
            data['parcela'] = parcela

    # Intravilan/Extravilan fallback from header
    if data['intravilan'] is None:
        if re.search(r'TEREN\s+Extravilan', full_text, re.IGNORECASE):
            data['intravilan'] = 'Extravilan'
        elif re.search(r'TEREN\s+Intravilan', full_text, re.IGNORECASE):
            data['intravilan'] = 'Intravilan'

    # Titular - B. Partea II. Proprietari
    # Find blocks: Intabulare/Inscrierea provizorie + PROPRIETATE ... until next Act or C. Partea
    prop_blocks = re.findall(
        r'(Intabulare|[IÎi]nscrierea\s+provizori[eă]),?\s+drept\s+de\s+PROPRIETATE(.*?)(?=\nAct\s|\nC\.\s+Partea|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )

    if prop_blocks:
        last_entry = prop_blocks[-1]
        tip_inscriere = last_entry[0].strip()
        block_text = last_entry[1]

        # Extract names from "N) NAME" pattern
        names = re.findall(r'\d+\)\s+(.+)', block_text)
        if names:
            cleaned = []
            for n in names:
                n = n.strip()
                n = re.sub(r'\b([A-ZĂÂÎȘȚ]{1,2})\.\s*', '', n)
                n = re.sub(r'^S\.?C\.?\s+', '', n)
                n = n.strip().rstrip(',').strip()
                if n:
                    cleaned.append(n)
            if cleaned:
                data['titular'] = ', '.join(cleaned)
                for nm in cleaned:
                    if re.search(r'NEIDENTIFICAT', nm, re.IGNORECASE):
                        data['titular'] = f"UAT (proprietar neidentificat)"
                    elif re.search(r'COMUNA\b', nm, re.IGNORECASE):
                        if 'provizori' in tip_inscriere.lower():
                            data['titular'] = f"UAT {nm} (înscriere provizorie)"
                        elif re.search(r'domeniu\s+privat', block_text, re.IGNORECASE):
                            data['titular'] = f"UAT {nm} - domeniu privat"
                        else:
                            data['titular'] = nm

    # If titular is neidentificat, check B3 for posesie faptică
    if data['titular'] and 'neidentificat' in data['titular'].lower():
        posesie_match = re.search(
            r'[Ss]e\s+noteaz[aă]\s+posesia\s+faptic[aă]\s+[iî]n\s+favoarea\s*\n?\s*\d+\)\s+(.+)',
            full_text, re.IGNORECASE
        )
        if posesie_match:
            posesor = posesie_match.group(1).strip()
            posesor = re.sub(r'\b([A-ZĂÂÎȘȚ]{1,2})\.\s*', '', posesor)
            data['titular'] = f"{posesor} (proprietar neidentificat)"

    # Sarcini - C. Partea III
    sarcini_section = re.search(
        r'C\.\s+Partea\s+III\.?\s*SARCINI.*?(?:NU\s+SUNT|Inscrieri\s+privind)(.*?)(?:Document\s+care|Pagina|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if sarcini_section:
        sarcini_text = sarcini_section.group(0)
        if re.search(r'NU\s+SUNT', sarcini_text, re.IGNORECASE):
            data['sarcini'] = []
        else:
            sarcini_items = []
            ipoteca_matches = re.findall(
                r'(?:Intabulare|Inscrierea).*?drept\s+de\s+IPOTEC[AĂ].*?\n.*?\d+\)\s+(.+)',
                sarcini_text, re.IGNORECASE
            )
            for im in ipoteca_matches:
                creditor = im.strip()
                creditor = re.sub(r'^S\.?C\.?\s+', '', creditor)
                sarcini_items.append(f"{creditor} (creditor ipotecar)")

            sechestru_matches = re.findall(
                r'sechestru\s+asigur[aă]tor.*?\n.*?\d+\)\s+(.+)',
                sarcini_text, re.IGNORECASE
            )
            for sm in sechestru_matches:
                sarcini_items.append(f"{sm.strip()} (sechestru asigurător)")

            if re.search(r'interdic[tț]ie', sarcini_text, re.IGNORECASE):
                sarcini_items.append("interdicție")
            if re.search(r'privilegiu', sarcini_text, re.IGNORECASE):
                sarcini_items.append("privilegiu")

            data['sarcini'] = sarcini_items

    return data


def extract_pad_data(pad_path):
    """Extract data from Plan de Amplasament si Delimitare (PAD) PDF."""
    data = {
        'suprafata_expropriata': None,
        'categorie': None,
        'intravilan': None,
        'tarla': None,
        'parcela': None,
        'constructii': [],
        'categorii_multiple': [],
    }

    full_text = extract_text_from_pdf(pad_path)
    if not full_text.strip():
        return data

    # Suprafata
    m2 = re.search(r'Suprafata\s+totala\s+masurata\s+a\s+imobilului\s*=?\s*([\d.]+)\s*mp', full_text, re.IGNORECASE)
    if m2:
        data['suprafata_expropriata'] = parse_suprafata(m2.group(1))

    if data['suprafata_expropriata'] is None:
        m = re.search(r'Suprafata\s+masurata\s+a\s+imobilului\s*\(mp\).*?\n.*?(\d+)', full_text, re.IGNORECASE)
        if m:
            data['suprafata_expropriata'] = int(m.group(1))

    if data['suprafata_expropriata'] is None:
        m3 = re.search(r'[=\s](\d{2,6})\s*mp', full_text)
        if m3:
            data['suprafata_expropriata'] = int(m3.group(1))
        else:
            m4 = re.search(r'Total\s+(\d+)', full_text)
            if m4:
                data['suprafata_expropriata'] = int(m4.group(1))

    # Tarla & Parcela - multiple patterns for different PDF formats
    # Pattern 1: "Tarla: 197/1, Parcela: 2768/1" (OCR format from header)
    m = re.search(r'Tarla:?\s*([\d/]+),?\s*Parcela:?\s*([\d/]+)', full_text, re.IGNORECASE)
    if m:
        data['tarla'] = m.group(1)
        data['parcela'] = m.group(2)

    # Pattern 2: "Tarla 29 Parcela 382/1"
    if not data['tarla']:
        m = re.search(r'Tarla\s+(\d+)\s+Parcela\s+([\d/]+)', full_text, re.IGNORECASE)
        if m:
            data['tarla'] = m.group(1)
            data['parcela'] = m.group(2)

    # Pattern 3: "Tarla 29, P 382/1"
    if not data['tarla']:
        m_lot = re.search(r'Tarla\s+(\d+),?\s*P(?:arcela)?\s+([\d/]+)', full_text, re.IGNORECASE)
        if m_lot:
            data['tarla'] = m_lot.group(1)
            if '/' in m_lot.group(2):
                data['parcela'] = m_lot.group(2)

    # Categorie from A. Date referitoare la teren
    teren_section = re.search(r'A\.\s*Date\s+referitoare\s+la\s+teren(.*?)(?:B\.\s*Date|Total)', full_text, re.IGNORECASE | re.DOTALL)
    if teren_section:
        teren_text = teren_section.group(1)
        lot_entries = re.findall(
            r'LOT\s+\d+\s*[-–]\s*'
            r'([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+?)\s+'
            r'(intravilan|extravilan)',
            teren_text, re.IGNORECASE
        )
        if lot_entries:
            data['categorie'] = normalize_cat(lot_entries[0][0].strip())
            data['intravilan'] = lot_entries[0][1].capitalize()
        else:
            cat_match = re.search(
                r'\d+\s+([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+?)\s+(\d+)\s+',
                teren_text
            )
            if cat_match:
                data['categorie'] = normalize_cat(cat_match.group(1))

        # Multiple categories
        all_cats = re.findall(
            r'(?:LOT\s+\d+\s*[-–]\s*)?'
            r'([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+?)\s+'
            r'(intravilan|extravilan)',
            teren_text, re.IGNORECASE
        )
        if len(all_cats) > 1:
            unique_cats = []
            seen = set()
            for cat_raw, pos in all_cats:
                cat = normalize_cat(cat_raw)
                key = (cat, pos.capitalize())
                if key not in seen:
                    seen.add(key)
                    unique_cats.append({'categorie': cat, 'intravilan': pos.capitalize()})
            if len(unique_cats) > 1:
                data['categorii_multiple'] = unique_cats

    # Intravilan/Extravilan - also from Mentiuni column in OCR
    if data['intravilan'] is None:
        # Look for "Extravilan" or "Intravilan" near LOT or Mentiuni
        if re.search(r'(?:Mentiuni|LOT\s*\d*).*?Extravilan', full_text, re.IGNORECASE | re.DOTALL):
            data['intravilan'] = 'Extravilan'
        elif re.search(r'(?:Mentiuni|LOT\s*\d*).*?Intravilan', full_text, re.IGNORECASE | re.DOTALL):
            data['intravilan'] = 'Intravilan'
        elif re.search(r'extravilan', full_text, re.IGNORECASE):
            data['intravilan'] = 'Extravilan'
        elif re.search(r'intravilan', full_text, re.IGNORECASE):
            data['intravilan'] = 'Intravilan'

    # Categorie from OCR - look for common category patterns near "Categoria de folosinta"
    if data['categorie'] is None:
        cat_match = re.search(r'Categori[ea]\s+de\s+folosint[aă]\s+.*?([Aa]rabil|[Ff][aăâ]ne[tț][aăe]|[Pp][aă][sș]une|[Nn]eproductiv|[Cc]ur[tț]i)', full_text, re.IGNORECASE | re.DOTALL)
        if cat_match:
            data['categorie'] = normalize_cat(cat_match.group(1))

    # Also try: header line "Nr. cadastral |Suprafata... Tarla: X, Parcela: Y,LOT 1" followed by category in Mentiuni
    if data['categorie'] is None:
        # OCR pattern: "1A" near LOT = arabil
        cat_codes = re.findall(r'\b(\d+)\s*([AFPNC])\b', full_text)
        for num, code in cat_codes:
            code_map = {'A': 'Arabil', 'F': 'Faneata', 'P': 'Pasune', 'N': 'Neproductiv', 'C': 'Curti constructii'}
            if code in code_map and int(num) > 100:  # likely "2500 A" = area + category
                continue
            if code in code_map:
                data['categorie'] = code_map[code]
                break

    # Constructii
    constr_section = re.search(
        r'B\.\s*Date\s+referitoare\s+la\s+construc[tț]ii(.*?)(?:Suprafata\s+totala|Total)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if constr_section:
        constr_text = constr_section.group(1)
        constr_entries = re.findall(r'(C\d+)\s+(\w+)\s+(\d+)', constr_text)
        for c in constr_entries:
            data['constructii'].append({
                'cod': c[0],
                'destinatie': c[1],
                'suprafata': int(c[2])
            })

    return data


def extract_memoriu_data(anexe_path):
    """Extract data from Memoriu Tehnic within ANEXE.pdf."""
    data = {
        'pozitii_hg': [],
        'suprafata': None,
        'categorie': None,
        'intravilan': None,
    }

    full_text = extract_text_from_pdf(anexe_path)
    if not full_text.strip():
        return data

    # Extract memoriu tehnic section
    memoriu_text = ""
    memoriu_match = re.search(r'MEMORIU\s+TEHNIC(.*?)(?:CALCULUL\s+ANALITIC|ANEXA\s+NR|$)', full_text, re.IGNORECASE | re.DOTALL)
    if memoriu_match:
        memoriu_text = memoriu_match.group(0)

    if not memoriu_text:
        memoriu_text = full_text  # fallback: use entire text

    poz_matches = re.findall(r'poz(?:i[tț]i[ae])?\s*\.?\s*(\d+)', memoriu_text, re.IGNORECASE)
    if poz_matches:
        data['pozitii_hg'] = [int(p) for p in poz_matches]

    m = re.search(r'suprafata\s+(?:de\s+)?(?:teren\s+)?(?:de\s+)?(\d+)\s*mp', memoriu_text, re.IGNORECASE)
    if m:
        data['suprafata'] = int(m.group(1))

    m2 = re.search(r'[iî]n\s+suprafata\s+de\s+([\d.]+)\s*mp', memoriu_text, re.IGNORECASE)
    if m2:
        sup_str = m2.group(1).replace('.', '')
        try:
            data['suprafata'] = int(sup_str)
        except ValueError:
            pass

    m = re.search(r'Categorie\s+de\s+folosit[aă]:\s*([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+)', memoriu_text, re.IGNORECASE)
    if m:
        data['categorie'] = normalize_cat(m.group(1))

    return data


def extract_incheiere_data(incheiere_path):
    """Extract data from INCHEIERE PDF (respingere or admitere)."""
    data = {
        'nr_dosar': None,
        'data_dosar': None,
        'tip': None,
    }

    full_text = extract_text_from_pdf(incheiere_path)
    if not full_text.strip():
        return data

    m = re.search(r'Incheiere\s+Nr\.?\s*(\d+)\s*/\s*(\d{2}[-./]\d{2}[-./]\d{4})', full_text, re.IGNORECASE)
    if m:
        data['nr_dosar'] = m.group(1)
        date_str = m.group(2)
        date_str = date_str.replace('-', '.').replace('/', '.')
        data['data_dosar'] = date_str

    if re.search(r'RESPINGERE', full_text, re.IGNORECASE):
        data['tip'] = 'RESPINGERE'
    elif re.search(r'ADMITERE|Admiterea\s+cererii|DISPUNE\s*\n\s*Admiterea', full_text, re.IGNORECASE):
        data['tip'] = 'ADMITERE'
    elif re.search(r'Admiterea', full_text):
        data['tip'] = 'ADMITERE'

    return data


# ── Scan complet PDF extraction ─────────────────────────────────────────
def extract_scan_complet_data(pdf_path):
    """Extract all data from a single scanned PDF containing full documentation.
    The PDF contains CF extract, PAD, memoriu tehnic, incheieri - all in one file."""
    data = {
        'cf_data': {},
        'pad_data': {},
        'memoriu_data': {},
        'incheiere_resp_data': {},
        'admitere_status': None,
    }

    full_text = extract_text_from_pdf(pdf_path)
    if not full_text.strip():
        return data

    # Try to split text into sections and extract data
    # CF section
    cf_section = re.search(
        r'(?:EXTRAS\s+DE\s+CARTE\s+FUNCIAR[AĂ]|EXTRASUL\s+DE\s+CARTE\s+FUNCIAR[AĂ])(.*?)(?:PLAN\s+DE\s+AMPLASAMENT|MEMORIU\s+TEHNIC|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if cf_section:
        cf_text = cf_section.group(0)
        data['cf_data'] = _extract_cf_from_text(cf_text)
    else:
        # Try extracting CF data from the whole text
        data['cf_data'] = _extract_cf_from_text(full_text)

    # PAD section
    pad_section = re.search(
        r'(PLAN\s+DE\s+AMPLASAMENT.*?)(?:MEMORIU\s+TEHNIC|INCHEIERE|CALCULUL\s+ANALITIC|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if pad_section:
        pad_text = pad_section.group(0)
        data['pad_data'] = _extract_pad_from_text(pad_text)

    # Memoriu tehnic section
    memoriu_section = re.search(
        r'(MEMORIU\s+TEHNIC.*?)(?:CALCULUL\s+ANALITIC|INCHEIERE|\Z)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if memoriu_section:
        mt_text = memoriu_section.group(1)
        m = re.search(r'Categorie\s+de\s+folosit[aă]:\s*([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+)', mt_text, re.IGNORECASE)
        if m:
            data['memoriu_data'] = {'categorie': normalize_cat(m.group(1))}

    # Incheiere
    incheiere_section = re.search(
        r'(INCHEIERE.*)',
        full_text, re.IGNORECASE | re.DOTALL
    )
    if incheiere_section:
        inch_text = incheiere_section.group(1)
        inch_data = {'nr_dosar': None, 'data_dosar': None, 'tip': None}
        m = re.search(r'Nr\.?\s*(\d+)\s*/\s*(\d{2}[-./]\d{2}[-./]\d{4})', inch_text)
        if m:
            inch_data['nr_dosar'] = m.group(1)
            inch_data['data_dosar'] = m.group(2).replace('-', '.').replace('/', '.')
        if re.search(r'RESPINGERE', inch_text, re.IGNORECASE):
            inch_data['tip'] = 'RESPINGERE'
        elif re.search(r'ADMITERE|Admiterea', inch_text, re.IGNORECASE):
            inch_data['tip'] = 'ADMITERE'
        data['incheiere_resp_data'] = inch_data

    # Admitere status
    sup_totala = data['cf_data'].get('suprafata_totala')
    sup_exp = data['pad_data'].get('suprafata_expropriata')
    if sup_totala and sup_exp and sup_totala == sup_exp:
        data['admitere_status'] = 'EXPROPRIAT INTEGRAL'
    elif data['incheiere_resp_data'].get('tip') == 'ADMITERE':
        data['admitere_status'] = 'ADMIS'
    elif sup_totala and sup_exp and sup_totala != sup_exp:
        data['admitere_status'] = 'LIPSĂ'
    else:
        data['admitere_status'] = 'EXPROPRIAT INTEGRAL'

    return data


def _extract_cf_from_text(text):
    """Extract CF data from a text block."""
    data = {
        'nr_cadastral': None, 'nr_cf': None, 'suprafata_totala': None,
        'categorie': None, 'intravilan': None, 'tarla': None,
        'parcela': None, 'titular': None, 'sarcini': [],
    }

    m = re.search(r'Carte\s+Funciar[aă]\s+Nr\.?\s*(\d+)', text, re.IGNORECASE)
    if m:
        data['nr_cf'] = int(m.group(1))
        data['nr_cadastral'] = int(m.group(1))

    m = re.search(r'A1\s+(\d+)\s+', text)
    if m:
        data['nr_cadastral'] = int(m.group(1))

    m = re.search(r'A1\s+\d+\s+([\d.,]+)', text)
    if m:
        data['suprafata_totala'] = parse_suprafata(m.group(1))

    teren_match = re.search(
        r'Date\s+referitoare\s+la\s+teren.*?'
        r'(\d+)\s+([a-zA-ZăâîșțĂÂÎȘȚşţ\-\s]+?)\s+'
        r'(DA|NU)\s+([\d.,]+)\s+(\d+|[-–])\s+([\d/]+|[-–])',
        text, re.IGNORECASE | re.DOTALL
    )
    if teren_match:
        data['categorie'] = normalize_cat(teren_match.group(2).strip())
        data['intravilan'] = 'Intravilan' if teren_match.group(3).upper() == 'DA' else 'Extravilan'
        tarla = teren_match.group(5).strip()
        if tarla and tarla not in ['-', '–']:
            data['tarla'] = tarla
        parcela = teren_match.group(6).strip()
        if parcela and parcela not in ['-', '–']:
            data['parcela'] = parcela

    if data['intravilan'] is None:
        if re.search(r'TEREN\s+Extravilan', text, re.IGNORECASE):
            data['intravilan'] = 'Extravilan'
        elif re.search(r'TEREN\s+Intravilan', text, re.IGNORECASE):
            data['intravilan'] = 'Intravilan'

    # Titular
    prop_blocks = re.findall(
        r'(Intabulare|[IÎi]nscrierea\s+provizori[eă]),?\s+drept\s+de\s+PROPRIETATE(.*?)(?=\nAct\s|\nC\.\s+Partea|\Z)',
        text, re.IGNORECASE | re.DOTALL
    )
    if prop_blocks:
        last_entry = prop_blocks[-1]
        tip_inscriere = last_entry[0].strip()
        block_text = last_entry[1]
        names = re.findall(r'\d+\)\s+(.+)', block_text)
        if names:
            cleaned = []
            for n in names:
                n = n.strip()
                n = re.sub(r'\b([A-ZĂÂÎȘȚ]{1,2})\.\s*', '', n)
                n = re.sub(r'^S\.?C\.?\s+', '', n)
                n = n.strip().rstrip(',').strip()
                if n:
                    cleaned.append(n)
            if cleaned:
                data['titular'] = ', '.join(cleaned)
                for nm in cleaned:
                    if re.search(r'NEIDENTIFICAT', nm, re.IGNORECASE):
                        data['titular'] = f"UAT (proprietar neidentificat)"
                    elif re.search(r'COMUNA\b', nm, re.IGNORECASE):
                        if 'provizori' in tip_inscriere.lower():
                            data['titular'] = f"UAT {nm} (înscriere provizorie)"
                        elif re.search(r'domeniu\s+privat', block_text, re.IGNORECASE):
                            data['titular'] = f"UAT {nm} - domeniu privat"
                        else:
                            data['titular'] = nm

    # Sarcini
    sarcini_section = re.search(
        r'C\.\s+Partea\s+III\.?\s*SARCINI.*?(?:NU\s+SUNT|Inscrieri\s+privind)(.*?)(?:Document\s+care|Pagina|\Z)',
        text, re.IGNORECASE | re.DOTALL
    )
    if sarcini_section:
        sarcini_text = sarcini_section.group(0)
        if not re.search(r'NU\s+SUNT', sarcini_text, re.IGNORECASE):
            sarcini_items = []
            ipoteca_matches = re.findall(
                r'(?:Intabulare|Inscrierea).*?drept\s+de\s+IPOTEC[AĂ].*?\n.*?\d+\)\s+(.+)',
                sarcini_text, re.IGNORECASE
            )
            for im in ipoteca_matches:
                creditor = re.sub(r'^S\.?C\.?\s+', '', im.strip())
                sarcini_items.append(f"{creditor} (creditor ipotecar)")
            sechestru_matches = re.findall(
                r'sechestru\s+asigur[aă]tor.*?\n.*?\d+\)\s+(.+)',
                sarcini_text, re.IGNORECASE
            )
            for sm in sechestru_matches:
                sarcini_items.append(f"{sm.strip()} (sechestru asigurător)")
            if re.search(r'interdic[tț]ie', sarcini_text, re.IGNORECASE):
                sarcini_items.append("interdicție")
            if re.search(r'privilegiu', sarcini_text, re.IGNORECASE):
                sarcini_items.append("privilegiu")
            data['sarcini'] = sarcini_items

    return data


def _extract_pad_from_text(text):
    """Extract PAD data from a text block."""
    data = {
        'suprafata_expropriata': None, 'categorie': None, 'intravilan': None,
        'tarla': None, 'parcela': None, 'constructii': [], 'categorii_multiple': [],
    }

    m2 = re.search(r'Suprafata\s+totala\s+masurata\s+a\s+imobilului\s*=?\s*([\d.]+)\s*mp', text, re.IGNORECASE)
    if m2:
        data['suprafata_expropriata'] = parse_suprafata(m2.group(1))

    if data['suprafata_expropriata'] is None:
        m = re.search(r'Suprafata\s+masurata\s+a\s+imobilului\s*\(mp\).*?\n.*?(\d+)', text, re.IGNORECASE)
        if m:
            data['suprafata_expropriata'] = int(m.group(1))

    if data['suprafata_expropriata'] is None:
        m3 = re.search(r'[=\s](\d{2,6})\s*mp', text)
        if m3:
            data['suprafata_expropriata'] = int(m3.group(1))

    m = re.search(r'Tarla\s+(\d+)\s+Parcela\s+([\d/]+)', text, re.IGNORECASE)
    if m:
        data['tarla'] = m.group(1)
        data['parcela'] = m.group(2)

    teren_section = re.search(r'A\.\s*Date\s+referitoare\s+la\s+teren(.*?)(?:B\.\s*Date|Total)', text, re.IGNORECASE | re.DOTALL)
    if teren_section:
        teren_text = teren_section.group(1)
        lot_entries = re.findall(
            r'LOT\s+\d+\s*[-–]\s*([A-Za-zăâîșțĂÂÎȘȚşţ\-\s]+?)\s+(intravilan|extravilan)',
            teren_text, re.IGNORECASE
        )
        if lot_entries:
            data['categorie'] = normalize_cat(lot_entries[0][0].strip())
            data['intravilan'] = lot_entries[0][1].capitalize()

    if data['intravilan'] is None:
        if re.search(r'extravilan', text, re.IGNORECASE):
            data['intravilan'] = 'Extravilan'
        elif re.search(r'intravilan', text, re.IGNORECASE):
            data['intravilan'] = 'Intravilan'

    constr_section = re.search(
        r'B\.\s*Date\s+referitoare\s+la\s+construc[tț]ii(.*?)(?:Suprafata\s+totala|Total)',
        text, re.IGNORECASE | re.DOTALL
    )
    if constr_section:
        constr_entries = re.findall(r'(C\d+)\s+(\w+)\s+(\d+)', constr_section.group(1))
        for c in constr_entries:
            data['constructii'].append({'cod': c[0], 'destinatie': c[1], 'suprafata': int(c[2])})

    return data


# ── File finding helpers ────────────────────────────────────────────────
def find_file_flexible(directory, patterns):
    """Find file matching any of the patterns (case-insensitive)."""
    if not os.path.isdir(directory):
        return None
    files = os.listdir(directory)
    for pattern in patterns:
        for f in files:
            if re.match(pattern, f, re.IGNORECASE):
                return os.path.join(directory, f)
    return None


def find_cf_pdf(folder_path):
    """Find CF extract PDF in folder."""
    files = os.listdir(folder_path)
    # Try EXTRAS.pdf first
    for f in files:
        if f.lower() in ['extras.pdf', 'extras cf.pdf']:
            return os.path.join(folder_path, f)
    for f in files:
        if f.lower().endswith('.pdf') and re.match(r'^(?:CF|EXTRAS)\s*', f, re.IGNORECASE):
            return os.path.join(folder_path, f)
    for f in files:
        if f.lower().endswith('.pdf') and re.match(r'^\d+\.pdf$', f, re.IGNORECASE):
            return os.path.join(folder_path, f)
    return None


def find_expropriat_folder(folder_path):
    """Find EXPROPRIAT folder (with flexible spelling)."""
    for item in os.listdir(folder_path):
        if os.path.isdir(os.path.join(folder_path, item)):
            if re.match(r'EXPRO[RP]*[IE]?AT', item, re.IGNORECASE):
                return os.path.join(folder_path, item)
    # Also try "Lot expropriat"
    for item in os.listdir(folder_path):
        if os.path.isdir(os.path.join(folder_path, item)):
            if re.search(r'expro', item, re.IGNORECASE):
                return os.path.join(folder_path, item)
    return None


def find_ramas_folder(folder_path):
    """Find RAMAS folder."""
    for item in os.listdir(folder_path):
        if os.path.isdir(os.path.join(folder_path, item)):
            if re.match(r'RAMAS', item, re.IGNORECASE):
                return os.path.join(folder_path, item)
    return None


def has_lot_subfolders(folder_path):
    """Check if folder has LOT subfolders."""
    for item in os.listdir(folder_path):
        if os.path.isdir(os.path.join(folder_path, item)):
            if re.match(r'LOT\s*\d+', item, re.IGNORECASE):
                return True
            if re.match(r'IE\s+\d+', item, re.IGNORECASE):
                return True
    return False


def get_lot_subfolders(folder_path):
    """Get sorted list of LOT subfolder paths."""
    lots = []
    for item in os.listdir(folder_path):
        full = os.path.join(folder_path, item)
        if os.path.isdir(full):
            lots.append(full)
    return sorted(lots)


def check_ramas_admitere(ramas_path):
    """Check RAMAS folder for ADMITERE incheiere."""
    if not ramas_path or not os.path.isdir(ramas_path):
        return None

    if has_lot_subfolders(ramas_path):
        for lot_path in get_lot_subfolders(ramas_path):
            inch_file = find_file_flexible(lot_path, [r'INCH[EI]+R[EI]*\.pdf', r'INCHEIERE.*\.pdf', r'Incheiere.*\.pdf'])
            if inch_file:
                inch_data = extract_incheiere_data(inch_file)
                if inch_data['tip'] == 'ADMITERE':
                    return 'ADMIS'
        return 'LIPSĂ'
    else:
        inch_file = find_file_flexible(ramas_path, [r'INCH[EI]+R[EI]*\.pdf', r'INCHEIERE.*\.pdf', r'Incheiere.*\.pdf'])
        if inch_file:
            inch_data = extract_incheiere_data(inch_file)
            if inch_data['tip'] == 'ADMITERE':
                return 'ADMIS'
            return 'LIPSĂ'
        return 'LIPSĂ'


# ── Folder-based extraction (HG 1119 style) ────────────────────────────
def extract_folder_data(folder_path):
    """Extract all data from a single position folder with EXPROPRIAT/RAMAS structure."""
    result = {
        'cf_data': {},
        'pad_data': {},
        'memoriu_data': {},
        'incheiere_resp_data': {},
        'admitere_status': None,
    }

    # CF extract - look in main folder
    cf_pdf = find_cf_pdf(folder_path)
    if cf_pdf:
        result['cf_data'] = extract_cf_data(cf_pdf)

    # EXPROPRIAT folder
    exp_folder = find_expropriat_folder(folder_path)
    if exp_folder:
        exp_has_lots = False
        exp_has_direct_pdf = False
        for item in os.listdir(exp_folder):
            item_path = os.path.join(exp_folder, item)
            if os.path.isdir(item_path) and re.match(r'LOT', item, re.IGNORECASE):
                exp_has_lots = True
            if item.upper().endswith('.PDF'):
                exp_has_direct_pdf = True

        if exp_has_lots and not exp_has_direct_pdf:
            lot_dirs = sorted([os.path.join(exp_folder, d) for d in os.listdir(exp_folder)
                              if os.path.isdir(os.path.join(exp_folder, d))])
            total_sup_exp = 0
            first_pad = None
            first_memoriu = None
            first_inch = None
            for lot_dir in lot_dirs:
                pad_file = find_file_flexible(lot_dir, [r'PAD\.pdf', r'PAD\s*.+\.pdf'])
                if pad_file:
                    pd = extract_pad_data(pad_file)
                    if pd.get('suprafata_expropriata'):
                        total_sup_exp += pd['suprafata_expropriata']
                    if first_pad is None:
                        first_pad = pd

                if first_memoriu is None:
                    anexe_file = find_file_flexible(lot_dir, [r'ANEX[AE]?[SE]?\.pdf', r'Anex[ae]?s?\.pdf'])
                    if anexe_file:
                        first_memoriu = extract_memoriu_data(anexe_file)

                if first_inch is None:
                    inch_file = find_file_flexible(lot_dir, [
                        r'INCH[EI]+R[EI]*\s*(?:respingere)?\.pdf',
                        r'Incheiere\s*respingere\.pdf',
                        r'INCH[EI]+R[EI]*\.pdf',
                        r'Incheiere\.pdf'
                    ])
                    if inch_file:
                        first_inch = extract_incheiere_data(inch_file)

            if first_pad:
                result['pad_data'] = first_pad
                result['pad_data']['suprafata_expropriata'] = total_sup_exp
            if first_memoriu:
                result['memoriu_data'] = first_memoriu
            if first_inch:
                result['incheiere_resp_data'] = first_inch
        else:
            pad_file = find_file_flexible(exp_folder, [r'PAD\.pdf', r'PAD\s*.+\.pdf'])
            if pad_file:
                result['pad_data'] = extract_pad_data(pad_file)

            anexe_file = find_file_flexible(exp_folder, [r'ANEX[AE]?[SE]?\.pdf', r'Anex[ae]?s?\.pdf'])
            if not anexe_file:
                # Check parent folder for ANEXE
                anexe_file = find_file_flexible(folder_path, [r'ANEX[AE]?[SE]?\.pdf', r'Anex[ae]?s?\.pdf'])
            if anexe_file:
                result['memoriu_data'] = extract_memoriu_data(anexe_file)

            inch_file = find_file_flexible(exp_folder, [
                r'INCH[EI]+R[EI]*\s*(?:respingere)?\.pdf',
                r'Incheiere\s*respingere\.pdf',
                r'INCH[EI]+R[EI]*\.pdf',
                r'Incheiere\.pdf'
            ])
            if inch_file:
                result['incheiere_resp_data'] = extract_incheiere_data(inch_file)

    # Also check for INCHEIERE in main folder
    if not result['incheiere_resp_data']:
        inch_file = find_file_flexible(folder_path, [
            r'INCH[EI]+R[EI]*\.pdf', r'INCHEIERE.*\.pdf', r'Incheiere.*\.pdf'
        ])
        if inch_file:
            result['incheiere_resp_data'] = extract_incheiere_data(inch_file)

    # RAMAS - admitere status
    ramas_folder = find_ramas_folder(folder_path)
    sup_totala = result['cf_data'].get('suprafata_totala')
    sup_exp = result['pad_data'].get('suprafata_expropriata')

    if sup_totala and sup_exp and sup_totala == sup_exp:
        result['admitere_status'] = 'EXPROPRIAT INTEGRAL'
    elif ramas_folder:
        result['admitere_status'] = check_ramas_admitere(ramas_folder)
    else:
        result['admitere_status'] = 'EXPROPRIAT INTEGRAL'

    return result


# ── Detection & scanning ────────────────────────────────────────────────
def detect_doc_type(extract_dir, depth=0):
    """Detect documentation type: 'scan_complet' or 'folder_structure'.
    Searches recursively through wrapper folders up to 5 levels deep.
    Returns (type_str, items_list)."""
    if depth > 5:
        return 'unknown', []

    try:
        items = os.listdir(extract_dir)
    except OSError:
        return 'unknown', []

    # Check for folders with numeric names (folder structure type)
    folder_items = []
    pdf_items = []
    for item in items:
        full = os.path.join(extract_dir, item)
        if os.path.isdir(full):
            if re.match(r'^\d', item):
                folder_items.append(full)
        elif item.lower().endswith('.pdf'):
            pdf_items.append(full)

    if folder_items:
        return 'folder_structure', folder_items
    elif pdf_items:
        return 'scan_complet', pdf_items
    else:
        # Check ALL subdirectories (ZIP/folder upload may have wrapper folders)
        for item in items:
            full = os.path.join(extract_dir, item)
            if os.path.isdir(full):
                sub_type, sub_items = detect_doc_type(full, depth + 1)
                if sub_items:
                    return sub_type, sub_items
        return 'unknown', []


def parse_folder_name(folder_name):
    """Parse folder name to extract position(s) and nr cadastral.
    Examples: '1234_55001', '587,605- 55001', '1234-55001'
    Returns (primary_pos, secondary_positions, nr_cad)."""
    m = re.match(r'^([\d,]+)\s*[-_]\s*(.+)$', folder_name)
    if m:
        pos_str = m.group(1)
        nr_cad = m.group(2).strip()
        positions = [int(p.strip()) for p in pos_str.split(',') if p.strip()]
        return positions[0], positions[1:], nr_cad
    # Just a number
    m = re.match(r'^(\d+)', folder_name)
    if m:
        return int(m.group(1)), [], None
    return None, [], None


def parse_pdf_name(pdf_name):
    """Parse PDF filename to extract position number(s).
    Examples: 'POZITIA 1234.pdf', 'poz 1234.pdf', '1234.pdf',
              '1 si 2.pdf', '163,164.pdf', '295 si 296.pdf'
    Returns (primary_pos, secondary_positions) or (None, [])."""
    name = os.path.splitext(pdf_name)[0].strip()

    # Pattern: "POZITIA 1234" or "poz 1234"
    m = re.match(r'^(?:POZITIA|POZ\.?|P\.?)\s*(\d+)', name, re.IGNORECASE)
    if m:
        return int(m.group(1)), []

    # Pattern: "1 si 2" or "295 si 296" (comasate with "si")
    m = re.match(r'^(\d+)\s+si\s+(\d+)$', name, re.IGNORECASE)
    if m:
        return int(m.group(1)), [int(m.group(2))]

    # Pattern: "163,164" (comasate with comma)
    m = re.match(r'^(\d+)\s*,\s*(\d+)$', name)
    if m:
        return int(m.group(1)), [int(m.group(2))]

    # Pattern: just a number "142"
    m = re.match(r'^(\d+)$', name)
    if m:
        return int(m.group(1)), []

    return None, []


def scan_and_extract(extract_dir):
    """Scan the extracted directory, detect type, and extract data for all positions.
    Returns list of dicts with position info and extracted data."""
    doc_type, items = detect_doc_type(extract_dir)
    results = []

    # Check for EXTRA HG subfolder
    extra_hg_items = []
    for item_path in items[:]:
        folder_name = os.path.basename(item_path)
        if re.match(r'EXTRA\s*HG', folder_name, re.IGNORECASE):
            # This is the EXTRA HG folder - process its contents separately
            if os.path.isdir(item_path):
                sub_type, sub_items = detect_doc_type(item_path)
                extra_hg_items.extend(sub_items)
            items.remove(item_path)

    if doc_type == 'folder_structure':
        for folder_path in sorted(items):
            folder_name = os.path.basename(folder_path)
            primary_pos, secondary_pos, nr_cad = parse_folder_name(folder_name)
            if primary_pos is None:
                continue
            try:
                data = extract_folder_data(folder_path)
                results.append({
                    'pozitie': primary_pos,
                    'secondary_positions': secondary_pos,
                    'folder_name': folder_name,
                    'doc_type': 'folder_structure',
                    'is_extra_hg': False,
                    'data': data,
                    'error': None,
                })
            except Exception as e:
                results.append({
                    'pozitie': primary_pos,
                    'secondary_positions': secondary_pos,
                    'folder_name': folder_name,
                    'doc_type': 'folder_structure',
                    'is_extra_hg': False,
                    'data': None,
                    'error': str(e),
                })

        # Process EXTRA HG folders
        for folder_path in sorted(extra_hg_items):
            folder_name = os.path.basename(folder_path)
            primary_pos, secondary_pos, nr_cad = parse_folder_name(folder_name)
            if primary_pos is None:
                continue
            try:
                data = extract_folder_data(folder_path) if os.path.isdir(folder_path) else extract_scan_complet_data(folder_path)
                results.append({
                    'pozitie': primary_pos,
                    'secondary_positions': secondary_pos,
                    'folder_name': f"[EXTRA HG] {folder_name}",
                    'doc_type': 'folder_structure',
                    'is_extra_hg': True,
                    'data': data,
                    'error': None,
                })
            except Exception as e:
                results.append({
                    'pozitie': primary_pos,
                    'secondary_positions': secondary_pos,
                    'folder_name': f"[EXTRA HG] {folder_name}",
                    'doc_type': 'folder_structure',
                    'is_extra_hg': True,
                    'data': None,
                    'error': str(e),
                })

    elif doc_type == 'scan_complet':
        for pdf_path in sorted(items):
            pdf_name = os.path.basename(pdf_path)
            poz, sec_positions = parse_pdf_name(pdf_name)
            if poz is None:
                continue
            try:
                data = extract_scan_complet_data(pdf_path)
                results.append({
                    'pozitie': poz,
                    'secondary_positions': sec_positions,
                    'folder_name': pdf_name,
                    'doc_type': 'scan_complet',
                    'is_extra_hg': False,
                    'data': data,
                    'error': None,
                })
            except Exception as e:
                results.append({
                    'pozitie': poz,
                    'secondary_positions': sec_positions,
                    'folder_name': pdf_name,
                    'doc_type': 'scan_complet',
                    'is_extra_hg': False,
                    'data': None,
                    'error': str(e),
                })

    return doc_type, results


# ── Excel column mapping ────────────────────────────────────────────────
def detect_column_mapping(ws):
    """Detect column mapping from the Excel header row.
    Returns dict mapping field names to column numbers."""
    mapping = {
        'pozitie_hg': 1,         # A - Pozitie HG
        'proprietar_hg': 4,      # D - Proprietar din HG
        'categorie_hg': 7,       # G - Categorie folosinta HG
        'pozitionare_hg': 8,     # H - Extravilan/Intravilan HG
        'suprafata_exp_hg': 15,  # O - Suprafata expropriata HG
        'observatii': 18,        # R - OBSERVATII
        'proprietar': 19,        # S - Proprietar doc cad
        'tarla': 20,             # T - Tarla
        'parcela': 21,           # U - Parcela
        'categorie': 22,         # V - Categorie folosinta
        'pozitionare': 23,       # W - Extravilan/Intravilan
        'nr_cadastral': 24,      # X - Nr cadastral
        'nr_topo': 25,           # Y - Nr topo
        'nr_cf': 26,             # Z - Nr CF
        'suprafata_totala': 27,  # AA - Suprafata totala
        'constructii': 28,       # AB - Constructii
        'suprafata_exp_constr': 29,  # AC
        'suprafata_exp': 30,     # AD - Suprafata expropriata doc cad
        'respingere': 31,        # AE - Nr/Data respingere
        'admitere': 32,          # AF - Admitere rest proprietate
        'obs_cadastru': 33,      # AG - Observatii cadastru
        'data_intro': 34,        # AH - Data introducerii
        'operator': 35,          # AI - Operator
    }

    # Try to detect from header row (row 1 or 2)
    for header_row in [1, 2]:
        for col in range(1, 50):
            val = sv(ws, header_row, col)
            if val is None:
                continue
            val_upper = str(val).upper().strip()
            if 'OBSERVATII' in val_upper and ('CF' in val_upper or 'DOC' in val_upper or 'CAD' in val_upper):
                mapping['obs_cadastru'] = col
            elif 'OBSERVATII' in val_upper and col > 15:
                mapping['observatii'] = col
            elif 'OPERATOR' in val_upper:
                mapping['operator'] = col
            elif 'DATA INTRODUCERII' in val_upper:
                mapping['data_intro'] = col

    return mapping


# ── Build preview data ──────────────────────────────────────────────────
def build_preview(extracted_results, excel_path):
    """Build preview data for display, comparing extracted data with HG values.
    Returns list of preview items."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[0]
    col_map = detect_column_mapping(ws)

    preview = []
    for item in extracted_results:
        poz = item['pozitie']
        data = item['data']
        p = {
            'pozitie': poz,
            'folder_name': item['folder_name'],
            'doc_type': item['doc_type'],
            'error': item['error'],
            'secondary_positions': item['secondary_positions'],
            'already_filled': False,
            'not_found': False,
        }

        if data is None:
            p['error'] = item['error'] or 'Eroare la extractie'
            preview.append(p)
            continue

        row = find_row(ws, poz)
        if row is None:
            p['not_found'] = True
            preview.append(p)
            continue

        if is_row_filled(ws, row):
            p['already_filled'] = True
            preview.append(p)
            continue

        cf = data.get('cf_data', {})
        pad = data.get('pad_data', {})
        memoriu = data.get('memoriu_data', {})

        # HG values
        hg_sup_exp = sv(ws, row, col_map['suprafata_exp_hg'])
        hg_cat = sv(ws, row, col_map['categorie_hg'])
        hg_pozitie = sv(ws, row, col_map['pozitionare_hg'])
        hg_proprietar = sv(ws, row, col_map['proprietar_hg'])

        # Doc cad values
        titular = cf.get('titular', '')
        sarcini = cf.get('sarcini', [])
        if sarcini:
            titular_display = titular + '\n' + '\n'.join(sarcini) if titular else '\n'.join(sarcini)
        else:
            titular_display = titular or ''

        doc_sup_exp = pad.get('suprafata_expropriata')
        doc_cat = pad.get('categorie') or cf.get('categorie')
        doc_pozitie = pad.get('intravilan') or cf.get('intravilan')

        # Comparisons
        diffs = []
        if not compare_numeric(hg_sup_exp, doc_sup_exp):
            diffs.append('SUPRAFAȚA EXPROPRIATĂ')
        if not compare_values(hg_cat, doc_cat):
            diffs.append('CATEGORIA DE FOLOSINȚĂ')
        if not compare_values(hg_pozitie, doc_pozitie):
            diffs.append('POZIȚIONAREA FAȚĂ DE LOCALITATE')

        obs_text = ', '.join(diffs) if diffs else 'OK'

        # Obs cadastru
        mt_cat = memoriu.get('categorie') if isinstance(memoriu, dict) else None
        pad_cat = pad.get('categorie')
        obs_cad = []
        if mt_cat and pad_cat and normalize_cat(mt_cat) != normalize_cat(pad_cat):
            obs_cad.append(f"categorie in MT: {mt_cat}, in PAD: {pad_cat}")
        cf_parcela = cf.get('parcela')
        pad_parcela = pad.get('parcela')
        if cf_parcela and pad_parcela and str(cf_parcela) != str(pad_parcela):
            obs_cad.append(f"parcela in CF: {cf_parcela}, in PAD: {pad_parcela}")

        p.update({
            'row': row,
            'hg_proprietar': hg_proprietar,
            'hg_sup_exp': hg_sup_exp,
            'hg_cat': hg_cat,
            'hg_pozitie': hg_pozitie,
            'titular': titular_display,
            'tarla': cf.get('tarla') or pad.get('tarla'),
            'parcela': cf.get('parcela') or pad.get('parcela'),
            'nr_cadastral': cf.get('nr_cadastral'),
            'nr_cf': cf.get('nr_cf'),
            'categorie': doc_cat,
            'pozitionare': doc_pozitie,
            'suprafata_totala': cf.get('suprafata_totala'),
            'suprafata_exp': doc_sup_exp,
            'constructii': pad.get('constructii', []),
            'respingere': None,
            'admitere': data.get('admitere_status'),
            'observatii': obs_text,
            'obs_cadastru': '; '.join(obs_cad) if obs_cad else None,
            'is_ok': obs_text == 'OK',
        })

        # Respingere - format: "nr/dd.mm.yy"
        inch = data.get('incheiere_resp_data', {})
        if isinstance(inch, dict) and inch.get('nr_dosar'):
            nr = inch['nr_dosar']
            data_d = inch.get('data_dosar', '')
            if data_d:
                date_parts = re.match(r'(\d{2})\.(\d{2})\.(\d{2,4})', data_d)
                if date_parts:
                    dd = date_parts.group(1)
                    mm = date_parts.group(2)
                    yy = date_parts.group(3)[-2:]
                    p['respingere'] = f"{nr}/{dd}.{mm}.{yy}"
                else:
                    p['respingere'] = nr
            else:
                p['respingere'] = nr

        preview.append(p)

    wb.close()
    return preview


# ── Write to Excel ──────────────────────────────────────────────────────
def write_to_excel(extracted_results, excel_path, output_path):
    """Write extracted data to Excel file. Returns summary dict."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[0]
    col_map = detect_column_mapping(ws)

    # Find or note EXTRA HG sheet
    ws_extra = None
    col_map_extra = None
    for sheet_name in wb.sheetnames:
        if re.search(r'EXTRA\s*HG', sheet_name, re.IGNORECASE):
            ws_extra = wb[sheet_name]
            col_map_extra = detect_column_mapping(ws_extra)
            break

    total_processed = 0
    total_ok = 0
    total_diff = 0
    total_skipped = 0
    errors = []

    for item in extracted_results:
        poz = item['pozitie']
        data = item['data']
        is_extra = item.get('is_extra_hg', False)

        # Use EXTRA HG sheet if applicable and available
        active_ws = ws_extra if (is_extra and ws_extra) else ws
        active_col_map = col_map_extra if (is_extra and col_map_extra) else col_map

        if data is None:
            errors.append({'pozitie': poz, 'error': item['error'] or 'No data'})
            continue

        row = find_row(active_ws, poz)
        if row is None:
            total_skipped += 1
            continue

        if is_row_filled(active_ws, row):
            total_skipped += 1
            continue

        cf = data.get('cf_data', {})
        pad = data.get('pad_data', {})
        memoriu = data.get('memoriu_data', {})
        inch = data.get('incheiere_resp_data', {})

        # HG reference values
        hg_sup_exp = sv(active_ws, row, active_col_map['suprafata_exp_hg'])
        hg_cat = sv(active_ws, row, active_col_map['categorie_hg'])
        hg_pozitie = sv(active_ws, row, active_col_map['pozitionare_hg'])

        # Values to write
        doc_sup_exp = pad.get('suprafata_expropriata')
        doc_cat = pad.get('categorie') or cf.get('categorie')
        doc_pozitie = pad.get('intravilan') or cf.get('intravilan')
        tarla = cf.get('tarla') or pad.get('tarla')
        parcela = cf.get('parcela') or pad.get('parcela')
        nr_cad = cf.get('nr_cadastral')
        nr_cf = cf.get('nr_cf')
        sup_totala = cf.get('suprafata_totala')
        constructii = pad.get('constructii', [])

        # Titular with sarcini
        titular_text = cf.get('titular', '')
        sarcini = cf.get('sarcini', [])
        if sarcini:
            titular_text = titular_text + '\n' + '\n'.join(sarcini) if titular_text else '\n'.join(sarcini)

        # Observations
        obs_items = []
        if not compare_numeric(hg_sup_exp, doc_sup_exp):
            obs_items.append('SUPRAFAȚA EXPROPRIATĂ')
        if not compare_values(hg_cat, doc_cat):
            obs_items.append('CATEGORIA DE FOLOSINȚĂ')
        if not compare_values(hg_pozitie, doc_pozitie):
            obs_items.append('POZIȚIONAREA FAȚĂ DE LOCALITATE')

        obs_text = ', '.join(obs_items) if obs_items else 'OK'
        obs_font = FONT_OBS_OK if obs_text == 'OK' else FONT_OBS_DIFF

        # Obs cadastru - diferente interne in documentatia cadastrala
        mt_cat = memoriu.get('categorie') if isinstance(memoriu, dict) else None
        pad_cat = pad.get('categorie')
        cf_cat = cf.get('categorie')
        obs_cad_items = []

        # MT vs PAD categorie
        if mt_cat and pad_cat and normalize_cat(mt_cat) != normalize_cat(pad_cat):
            obs_cad_items.append(f"categorie de folosinta in MT este {mt_cat}, iar in PAD este {pad_cat}")

        # CF vs PAD categorie
        if cf_cat and pad_cat and normalize_cat(cf_cat) != normalize_cat(pad_cat):
            obs_cad_items.append(f"categorie de folosinta in extrasul CF este {cf_cat}, iar in PAD este {pad_cat}")

        # CF vs PAD parcela
        cf_parcela = cf.get('parcela')
        pad_parcela = pad.get('parcela')
        if cf_parcela and pad_parcela and str(cf_parcela) != str(pad_parcela):
            clean_cf = str(cf_parcela).replace('/', '')
            clean_pad = str(pad_parcela).replace('/', '')
            if clean_cf != clean_pad:
                obs_cad_items.append(f"parcela în extrasul CF este {cf_parcela}, iar în PAD este {pad_parcela}")

        # CF vs PAD tarla
        cf_tarla = cf.get('tarla')
        pad_tarla = pad.get('tarla')
        if cf_tarla and pad_tarla and str(cf_tarla) != str(pad_tarla):
            obs_cad_items.append(f"tarla în extrasul CF este {cf_tarla}, iar în PAD este {pad_tarla}")

        # CF vs PAD intravilan/extravilan
        cf_pozitie = cf.get('intravilan')
        pad_pozitie = pad.get('intravilan')
        if cf_pozitie and pad_pozitie and cf_pozitie.lower() != pad_pozitie.lower():
            obs_cad_items.append(f"pozitionare in extrasul CF este {cf_pozitie}, iar in PAD este {pad_pozitie}")
        obs_cad_text = '; '.join(obs_cad_items) if obs_cad_items else None

        # Respingere - format: "nr/dd.mm.yy"
        resp_text = None
        if isinstance(inch, dict) and inch.get('nr_dosar'):
            nr = inch['nr_dosar']
            data_d = inch.get('data_dosar', '')
            if data_d:
                # Convert dd.mm.yyyy to dd.mm.yy
                date_parts = re.match(r'(\d{2})\.(\d{2})\.(\d{2,4})', data_d)
                if date_parts:
                    dd = date_parts.group(1)
                    mm = date_parts.group(2)
                    yy = date_parts.group(3)[-2:]  # Last 2 digits
                    resp_text = f"{nr}/{dd}.{mm}.{yy}"
                else:
                    resp_text = nr
            else:
                resp_text = nr

        # Admitere
        admitere = data.get('admitere_status')

        # Categorii multiple din PAD
        categorii_multiple = pad.get('categorii_multiple', [])

        # Write main row columns
        c = active_col_map
        aw = active_ws
        ss(aw, row, c['observatii'], obs_text, obs_font)
        ss(aw, row, c['proprietar'], titular_text or None, FONT_DEFAULT)
        ss(aw, row, c['tarla'], tarla or ' - ', FONT_DEFAULT)
        ss(aw, row, c['parcela'], parcela or ' - ', FONT_DEFAULT)
        ss(aw, row, c['categorie'], doc_cat, FONT_DEFAULT)
        ss(aw, row, c['pozitionare'], doc_pozitie, FONT_DEFAULT)
        ss(aw, row, c['nr_cadastral'], nr_cad, FONT_DEFAULT)
        ss(aw, row, c['nr_topo'], ' - ', FONT_DEFAULT)
        ss(aw, row, c['nr_cf'], nr_cf, FONT_DEFAULT)
        ss(aw, row, c['suprafata_totala'], sup_totala, FONT_DEFAULT)

        # Constructii - rând separat per construcție
        if constructii:
            # First construction on main row
            first_constr = constructii[0]
            constr_text = f"{first_constr['cod']} {first_constr.get('destinatie', '')} {first_constr['suprafata']} mp"
            ss(aw, row, c['constructii'], constr_text, FONT_DEFAULT)
            ss(aw, row, c['suprafata_exp_constr'], first_constr['suprafata'], FONT_DEFAULT)

            # Additional constructions: insert rows below
            extra_rows_inserted = 0
            for ci in range(1, len(constructii)):
                constr = constructii[ci]
                insert_at = row + 1 + extra_rows_inserted
                aw.insert_rows(insert_at)
                constr_text_extra = f"{constr['cod']} {constr.get('destinatie', '')} {constr['suprafata']} mp"
                ss(aw, insert_at, c['constructii'], constr_text_extra, FONT_DEFAULT)
                ss(aw, insert_at, c['suprafata_exp_constr'], constr['suprafata'], FONT_DEFAULT)
                extra_rows_inserted += 1
        else:
            ss(aw, row, c['constructii'], ' - ', FONT_DEFAULT)
            ss(aw, row, c['suprafata_exp_constr'], ' - ', FONT_DEFAULT)

        ss(aw, row, c['suprafata_exp'], doc_sup_exp, FONT_SUP_EXP)
        ss(aw, row, c['respingere'], resp_text, FONT_DEFAULT)
        ss(aw, row, c['admitere'], admitere, FONT_DEFAULT)
        if obs_cad_text:
            ss(aw, row, c['obs_cadastru'], obs_cad_text, FONT_DEFAULT)
        ss(aw, row, c['data_intro'], datetime.now().strftime('%d.%m.%Y'), FONT_DEFAULT)
        ss(aw, row, c['operator'], 'Claude', FONT_DEFAULT)

        # Categorii multiple - inserare rânduri suplimentare sub poziția de HG
        if len(categorii_multiple) > 1:
            for ci in range(1, len(categorii_multiple)):
                cat_extra = categorii_multiple[ci]
                insert_at = row + 1
                aw.insert_rows(insert_at)
                ss(aw, insert_at, c['categorie'], cat_extra.get('categorie'), FONT_DEFAULT)
                ss(aw, insert_at, c['pozitionare'], cat_extra.get('intravilan'), FONT_DEFAULT)

        # Color row - verde=OK, rosu=diferente
        fill = FILL_GREEN if obs_text == 'OK' else FILL_RED
        color_row(aw, row, fill)

        total_processed += 1
        if obs_text == 'OK':
            total_ok += 1
        else:
            total_diff += 1

        # Handle secondary (comasated) positions
        for sec_pos in item.get('secondary_positions', []):
            sec_row = find_row(aw, sec_pos)
            if sec_row and not is_row_filled(aw, sec_row):
                ss(aw, sec_row, c['observatii'], f'comasat cu poziția {poz}', FONT_DEFAULT)
                ss(aw, sec_row, c['data_intro'], datetime.now().strftime('%d.%m.%Y'), FONT_DEFAULT)
                ss(aw, sec_row, c['operator'], 'Claude', FONT_DEFAULT)
                color_row(aw, sec_row, fill)

    wb.save(output_path)
    wb.close()

    return {
        'total_processed': total_processed,
        'total_ok': total_ok,
        'total_diff': total_diff,
        'total_skipped': total_skipped,
        'errors': errors,
    }


# ── ZIP extraction ──────────────────────────────────────────────────────
def extract_zip(zip_path, extract_dir):
    """Extract ZIP file to directory."""
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(extract_dir)


def extract_uploaded_folder(files, extract_dir):
    """Save uploaded files preserving folder structure.
    Flask sends webkitRelativePath in the filename."""
    for f in files:
        if not f.filename:
            continue
        # Preserve relative path from webkitdirectory upload
        rel_path = f.filename.replace('\\', '/')
        full_path = os.path.join(extract_dir, rel_path)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        f.save(full_path)
