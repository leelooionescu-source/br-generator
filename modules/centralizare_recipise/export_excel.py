"""
export_excel.py — Export centralizare recipise in Excel cu sheet-uri per HG

Parseaza corpul emailurilor pentru a extrage tabelele cu recipise:
  Header: Proiect (HG xxx/yyyy) :BR. x.x/dd.mm.yyyy
  Coloane: SUMA | BENEFICIAR | DATA + NR_RECIPISA
"""

import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from modules.centralizare_recipise.database import get_all_emails, BASE_DIR

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color='0D9488')
DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Pattern pentru header sectiune: Proiect (HG xxx/yyyy) :BR. x.x/dd.mm.yyyy
SECTION_PATTERN = re.compile(
    r'(.+?)\s*\(HG\s*(\d+[/-]\d{4})\)\s*:?\s*'
    r'(B[RP])\.?\s*(\d+(?:\.\d+)?)\s*[/-]?\s*(\d{2}[./]\d{2}[./]\d{4})?',
    re.IGNORECASE
)

# Pattern pentru suma (numere cu virgula sau punct ca separator mii)
SUMA_PATTERN = re.compile(r'^[\d,\.]+\.\d{2}$|^[\d\.]+,\d{2}$')


def parse_email_body(body):
    """Parseaza corpul emailului si extrage sectiunile cu recipise.

    Returneaza lista de dict-uri:
    [{
        'proiect': 'Autostrada de centura Bucuresti (Nord)',
        'hg': '681/2024',
        'br': 'BR 10.1',
        'data_br': '18.03.2026',
        'rows': [
            {'suma': '27,169.81', 'beneficiar': 'DRAGAN DANIEL...', 'nr_recipisa': '212228192/1', 'data_consemnare': '25/03/2026'},
        ]
    }]
    """
    if not body:
        return []

    lines = [l.strip() for l in body.split('\n')]
    sections = []
    current_section = None
    data_consemnare = ''
    collecting_rows = False

    for line in lines:
        if not line:
            continue

        # Check for section header
        m = SECTION_PATTERN.search(line)
        if m:
            current_section = {
                'proiect': m.group(1).strip(),
                'hg': m.group(2).replace(' ', ''),
                'br': f"{m.group(3).upper()} {m.group(4)}",
                'data_br': m.group(5) or '',
                'rows': [],
            }
            sections.append(current_section)
            collecting_rows = False
            data_consemnare = ''
            continue

        if current_section is None:
            continue

        # Check for table header (SUMA / BENEFICIAR / date)
        if 'SUMA' in line.upper() and 'BENEFICIAR' in line.upper():
            collecting_rows = True
            continue

        # Check if line is a date header (e.g. "25/03/2026")
        date_match = re.match(r'^(\d{2}[/\.]\d{2}[/\.]\d{4})$', line)
        if date_match:
            data_consemnare = date_match.group(1)
            collecting_rows = True
            continue

        if not collecting_rows:
            continue

        # Try to parse as data row: SUMA then BENEFICIAR then NR_RECIPISA
        # The body text has them on separate lines after stripping
        # Check if this line looks like a sum
        clean_line = line.replace('\t', '').strip()
        if SUMA_PATTERN.match(clean_line):
            # This is a sum - start a new row
            current_section['rows'].append({
                'suma': clean_line,
                'beneficiar': '',
                'nr_recipisa': '',
                'data_consemnare': data_consemnare,
            })
        elif current_section['rows'] and not current_section['rows'][-1]['beneficiar']:
            # This is the beneficiar name
            current_section['rows'][-1]['beneficiar'] = clean_line
        elif current_section['rows'] and not current_section['rows'][-1]['nr_recipisa']:
            # This is the nr recipisa
            current_section['rows'][-1]['nr_recipisa'] = clean_line

    return sections


def _apply_header(ws, columns, row=1):
    """Aplica header cu stil pe un sheet."""
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def generate_excel_report():
    """Genereaza fisier Excel cu sheet-uri per HG cu datele recipiselor.
    Returneaza calea fisierului."""
    wb = Workbook()
    wb.remove(wb.active)

    all_emails = get_all_emails()

    # Parseaza toate emailurile si grupeaza per HG
    hg_data = defaultdict(list)  # hg -> list of sections
    all_sections = []

    for email in all_emails:
        sections = parse_email_body(email.get('body', ''))
        for sec in sections:
            sec['expeditor'] = email.get('sender', '')
            sec['data_email'] = email.get('received_date', '')[:10] if email.get('received_date') else ''
            hg_key = sec['hg']
            hg_data[hg_key].append(sec)
            all_sections.append(sec)

    columns = [
        ('Nr.', 6),
        ('Proiect', 35),
        ('HG', 14),
        ('BR/BP', 12),
        ('Data BR', 14),
        ('Suma', 16),
        ('Beneficiar', 40),
        ('Nr. Recipisa', 18),
        ('Data Consemnare', 16),
        ('Expeditor', 28),
        ('Data Email', 14),
    ]

    columns_per_hg = [
        ('Nr.', 6),
        ('BR/BP', 12),
        ('Data BR', 14),
        ('Suma', 16),
        ('Beneficiar', 40),
        ('Nr. Recipisa', 18),
        ('Data Consemnare', 16),
        ('Expeditor', 28),
        ('Data Email', 14),
    ]

    # Sheet TOTAL
    ws_total = wb.create_sheet(title='TOTAL', index=0)
    _apply_header(ws_total, columns)
    row_num = 2
    for sec in all_sections:
        for r in sec['rows']:
            vals = [
                row_num - 1,
                sec['proiect'],
                sec['hg'],
                sec['br'],
                sec['data_br'],
                r['suma'],
                r['beneficiar'],
                r['nr_recipisa'],
                r['data_consemnare'],
                sec['expeditor'],
                sec['data_email'],
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws_total.cell(row=row_num, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_idx in (1, 3, 4, 5, 9, 11):
                    cell.alignment = CENTER_ALIGN
                elif col_idx == 6:
                    cell.alignment = RIGHT_ALIGN
                else:
                    cell.alignment = DATA_ALIGN
            row_num += 1

    ws_total.freeze_panes = 'A2'
    if row_num > 2:
        ws_total.auto_filter.ref = f"A1:K{row_num - 1}"

    # Sheet per HG
    for hg in sorted(hg_data.keys()):
        if not hg:
            continue
        sheet_name = f"HG {hg}".replace('/', '-')[:31]
        ws = wb.create_sheet(title=sheet_name)
        _apply_header(ws, columns_per_hg)

        row_num = 2
        sections = hg_data[hg]
        # Sorteaza dupa BR
        sections.sort(key=lambda s: s['br'])

        for sec in sections:
            for r in sec['rows']:
                vals = [
                    row_num - 1,
                    sec['br'],
                    sec['data_br'],
                    r['suma'],
                    r['beneficiar'],
                    r['nr_recipisa'],
                    r['data_consemnare'],
                    sec['expeditor'],
                    sec['data_email'],
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    if col_idx in (1, 2, 3, 7, 9):
                        cell.alignment = CENTER_ALIGN
                    elif col_idx == 4:
                        cell.alignment = RIGHT_ALIGN
                    else:
                        cell.alignment = DATA_ALIGN
                row_num += 1

        ws.freeze_panes = 'A2'
        if row_num > 2:
            ws.auto_filter.ref = f"A1:I{row_num - 1}"

    # Salveaza
    output_dir = os.path.join(BASE_DIR, 'data')
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, 'Centralizare Recipise.xlsx')
    wb.save(filepath)
    return filepath
