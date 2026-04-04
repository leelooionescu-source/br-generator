"""
process_master.py — Logica de procesare MASTER COMISII

Etapa 1: Import date din surse Excel -> MASTER
Etapa 2: Completare coloane 30-38 (HG, decizie, membri comisie)
Etapa 3: Generare Word prin mail merge MERGEFIELD
"""

import os
import re
import math
import shutil
import zipfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy


# ─── Formatare standard MASTER ────────────────────────────────────────────────
MASTER_FONT = Font(name='Trebuchet MS', size=10, bold=True)
MASTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
MASTER_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _apply_master_style(cell):
    cell.font = copy(MASTER_FONT)
    cell.alignment = copy(MASTER_ALIGN)
    cell.border = copy(MASTER_BORDER)


# ─── Mapari coloane ───────────────────────────────────────────────────────────
# Format A: col sursa (0-based) -> col MASTER (1-based)
SRC_A_TO_MASTER = {
    1: 2, 2: 3, 3: 4, 5: 12, 6: 13, 7: 14, 8: 15,
    9: 16, 10: 17, 11: 18, 12: 19, 13: 20,
}

# Format B: col sursa (1-based, ca index 0-based) -> col MASTER (1-based)
SRC_B_TO_MASTER = {
    0: 2, 1: 3, 2: 4, 4: 12, 5: 13, 6: 14, 7: 15,
    8: 16, 9: 17, 10: 18, 11: 19, 14: 20,
}

# MERGEFIELD name -> coloana MASTER (1-based)
FIELD_COL = {
    'Nr_PV_si_Hot': 1, 'POZITIE_HG': 2, 'JUDET': 3, 'UAT': 4,
    'NUME1': 6, 'ADRESA1': 7, 'NUME2': 8, 'ADRESA2': 9,
    'Tarla': 12, 'Parcel\u0103': 13, 'Nr_cadastral': 14, 'Nr_CF': 15,
    'Suprafata_expropriata_teren_1': 19, 'Valoare_despagubiri__teren_1': 20,
    'Suprafata_expropriata_teren_2': 21, 'Valoare_despagubiri__teren_2': 22,
    'Suprafata_constructie_1': 23, 'Valoare_despagubiri_constructie_1': 24,
    'Suprafata_constructie_2': 25, 'Valoare_despagubiri_constructie_2': 26,
    'Suprafata_constructie_3': 27, 'Valoare_despagubiri_constructie_3': 28,
    'HG': 29, 'Text_complet_HG': 30,
    'Decizie_de_expropriere': 32, 'Decizie_de_numire_comisie': 33,
    'MEMBRU1': 34, 'MEMBRU2': 35, 'MEMBRU3': 36, 'MEMBRU4': 37, 'MEMBRU5': 38,
    'Text_pentru_LCA_convocare': 39, 'DATA': 41,
}


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 1: Import date
# ═══════════════════════════════════════════════════════════════════════════════

def detect_format(path):
    """Detecteaza formatul fisierului sursa.
    Returns 'A' (Tabel afisare comisie) sau 'B' (SITUATIE TOTAL)."""
    df = pd.read_excel(path, header=None, nrows=5)
    # Format B: coloana 0 = Pozitie HG (numeric direct din randul 2+)
    # Format A: coloana 0 = Nr. Crt, coloana 1 = Pozitie HG
    # Verificam daca exista o coloana cu "OBSERVATII" in header-uri (Format B)
    for i in range(min(3, len(df))):
        for j in range(df.shape[1]):
            val = str(df.iloc[i, j]).upper() if pd.notna(df.iloc[i, j]) else ''
            if 'OBSERV' in val:
                return 'B'
    # Verificam daca numele fisierului indica formatul
    basename = os.path.basename(path).upper()
    if 'SITUATIE' in basename or 'SITUATIE TOTAL' in basename:
        return 'B'
    if 'TABEL' in basename or 'COMISIE' in basename:
        return 'A'
    return 'A'  # default


def read_source_a(path):
    """Citeste Format A (Tabel afisare comisie). Returneaza lista de dict-uri."""
    raw = pd.read_excel(path, header=None)
    data = raw.iloc[2:].copy()
    data.columns = range(len(data.columns))
    # Pastreaza doar randuri cu Nr. Crt numeric (coloana 0)
    data = data[pd.to_numeric(data[0], errors='coerce').notna()].copy()
    data = data.reset_index(drop=True)

    rows = []
    for _, row in data.iterrows():
        entry = {}
        for src_col, master_col in SRC_A_TO_MASTER.items():
            val = row[src_col] if src_col < len(row) else None
            if isinstance(val, float) and (math.isnan(val) or pd.isna(val)):
                val = None
            entry[master_col] = val

        # Numes (split dupa virgula -> cols 6, 8, 10)
        numes_raw = str(row[4]).strip() if pd.notna(row[4]) else ''
        numes = [n.strip() for n in numes_raw.split(',') if n.strip()]
        entry[6] = numes[0] if len(numes) > 0 else None
        entry[8] = numes[1] if len(numes) > 1 else None
        entry[10] = numes[2] if len(numes) > 2 else None

        rows.append(entry)
    return rows


def read_source_b(path):
    """Citeste Format B (SITUATIE TOTAL). Filtreaza doar OBSERVATII=OK.
    Returneaza lista de dict-uri."""
    raw = pd.read_excel(path, header=None)

    # Gaseste coloana OBSERVATII
    obs_col = None
    header_row = 0
    for i in range(min(3, len(raw))):
        for j in range(raw.shape[1]):
            val = str(raw.iloc[i, j]).upper() if pd.notna(raw.iloc[i, j]) else ''
            if 'OBSERV' in val:
                obs_col = j
                header_row = i
                break
        if obs_col is not None:
            break

    data = raw.iloc[header_row + 1:].copy()
    data.columns = range(len(data.columns))
    # Pastreaza doar randuri cu Pozitie HG numerica (coloana 0)
    data = data[pd.to_numeric(data[0], errors='coerce').notna()].copy()
    data = data.reset_index(drop=True)

    rows = []
    filtered_count = 0
    for _, row in data.iterrows():
        # Filtreaza dupa OBSERVATII daca exista coloana
        if obs_col is not None:
            obs_val = str(row[obs_col]).strip().upper() if pd.notna(row[obs_col]) else ''
            if obs_val != 'OK':
                filtered_count += 1
                continue

        entry = {}
        for src_col, master_col in SRC_B_TO_MASTER.items():
            val = row[src_col] if src_col < len(row) else None
            if isinstance(val, float) and (math.isnan(val) or pd.isna(val)):
                val = None
            entry[master_col] = val

        # Numes (split dupa virgula -> cols 6, 8, 10)
        numes_raw = str(row[3]).strip() if pd.notna(row[3]) else ''
        numes = [n.strip() for n in numes_raw.split(',') if n.strip()]
        entry[6] = numes[0] if len(numes) > 0 else None
        entry[8] = numes[1] if len(numes) > 1 else None
        entry[10] = numes[2] if len(numes) > 2 else None

        rows.append(entry)

    return rows, filtered_count


def preview_import(source_path, fmt):
    """Previzualizeaza datele care vor fi importate. Returneaza statistici + primele randuri."""
    if fmt == 'A':
        rows = read_source_a(source_path)
        filtered = 0
    else:
        rows, filtered = read_source_b(source_path)

    # Extrage UAT-uri unice
    uats = sorted(set(str(r.get(4, '')).strip() for r in rows if r.get(4)))

    preview_rows = []
    for r in rows[:20]:
        preview_rows.append({
            'poz_hg': r.get(2, ''),
            'judet': r.get(3, ''),
            'uat': r.get(4, ''),
            'nume1': r.get(6, ''),
            'nr_cadastral': r.get(14, ''),
            'suprafata': r.get(19, ''),
            'valoare': r.get(20, ''),
        })

    return {
        'total': len(rows),
        'filtered': filtered,
        'uats': uats,
        'preview': preview_rows,
        'format': fmt,
    }


def import_to_master(master_path, source_path, fmt, hg_nr, data_sedinta, output_path):
    """Importa datele din sursa in MASTER. Salveaza in output_path."""
    # Citeste sursa
    if fmt == 'A':
        source_rows = read_source_a(source_path)
    else:
        source_rows, _ = read_source_b(source_path)

    if not source_rows:
        raise ValueError('Nu s-au gasit date de importat din fisierul sursa.')

    # Copie MASTER la output
    shutil.copy2(master_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Gaseste ultimul rand cu date
    last_row = 1
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, 45)):
            last_row = r
            break

    # Primul numar secvential disponibil
    existing_max = 0
    for r in range(2, last_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, (int, float)) and not pd.isna(v):
            existing_max = max(existing_max, int(v))

    pv_counter = existing_max + 1
    start_row = last_row + 1 if last_row > 1 else 2

    for idx, entry in enumerate(source_rows):
        r = start_row + idx
        # Nr. PV si Hot (secvential)
        cell = ws.cell(r, 1, pv_counter)
        _apply_master_style(cell)
        pv_counter += 1

        # Coloane mapate
        for master_col, val in entry.items():
            cell = ws.cell(r, master_col, val)
            _apply_master_style(cell)

        # HG si DATA
        cell = ws.cell(r, 29, hg_nr)
        _apply_master_style(cell)
        if data_sedinta:
            cell = ws.cell(r, 41, data_sedinta)
            _apply_master_style(cell)

        # Aplica formatare pe toate coloanele
        for c in range(1, 45):
            cell = ws.cell(r, c)
            _apply_master_style(cell)

    wb.save(output_path)
    return {
        'imported': len(source_rows),
        'start_row': start_row,
        'end_row': start_row + len(source_rows) - 1,
        'pv_start': existing_max + 1,
        'pv_end': pv_counter - 1,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 2: Completare coloane 30-38
# ═══════════════════════════════════════════════════════════════════════════════

def get_uats_from_master(master_path):
    """Extrage UAT-uri unice din MASTER (col 4)."""
    wb = load_workbook(master_path, read_only=True)
    ws = wb.active
    uats = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 4).value
        if val:
            uats.add(str(val).strip())
    wb.close()
    return sorted(uats)


def get_master_stats(master_path):
    """Statistici din MASTER: total randuri, randuri per UAT, coloane completate."""
    wb = load_workbook(master_path, read_only=True)
    ws = wb.active
    total = 0
    uat_counts = {}
    cols_filled = {30: 0, 31: 0, 32: 0, 33: 0, 34: 0}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        total += 1
        uat = str(ws.cell(r, 4).value or '').strip()
        uat_counts[uat] = uat_counts.get(uat, 0) + 1
        for c in cols_filled:
            if ws.cell(r, c).value:
                cols_filled[c] += 1
    wb.close()
    return {
        'total': total,
        'uat_counts': uat_counts,
        'cols_filled': cols_filled,
    }


def update_cols_30_38(master_path, config, output_path):
    """Actualizeaza coloanele 30-38 din MASTER.

    config = {
        'fixed': {
            30: 'Text complet HG...',
            31: 'Antet...',
            32: 'Decizie expropriere...',
        },
        'per_uat': {
            'MARGINA': {
                33: 'Decizie numire comisie...',
                34: 'Membru1', 35: 'Membru2', 36: 'Membru3',
                37: 'Membru4', 38: 'Membru5',
            },
            ...
        }
    }
    """
    shutil.copy2(master_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    updated = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue

        uat = str(ws.cell(r, 4).value or '').strip().upper()

        # Valori fixe
        for col, val in config.get('fixed', {}).items():
            col = int(col)
            if val:
                cell = ws.cell(r, col, val)
                _apply_master_style(cell)
                if col == 31:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True)

        # Valori per UAT
        for uat_key, uat_vals in config.get('per_uat', {}).items():
            if uat_key.upper() in uat or uat in uat_key.upper():
                for col, val in uat_vals.items():
                    col = int(col)
                    if val:
                        cell = ws.cell(r, col, val)
                        _apply_master_style(cell)
                break

        updated += 1

    wb.save(output_path)
    return updated


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 3: Generare Word mail merge
# ═══════════════════════════════════════════════════════════════════════════════

def _get_val(ws, r, col):
    """Citeste celula din MASTER, returneaza string."""
    v = ws.cell(r, col).value
    if v is None:
        return ''
    if isinstance(v, float):
        if math.isnan(v):
            return ''
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v)


def _xml_escape(s):
    return (s.replace('&', '&amp;')
             .replace('<', '&lt;')
             .replace('>', '&gt;')
             .replace('"', '&quot;'))


def _replace_field(xml, field_name, value):
    """Inlocuieste valoarea afisata a unui MERGEFIELD in XML-ul Word."""
    val = _xml_escape(value)
    pattern = (
        r'(<w:instrText[^>]*>[^<]*MERGEFIELD\s+' + re.escape(field_name) + r'\s[^<]*</w:instrText>'
        r'.*?<w:fldChar[^>]+fldCharType="separate"[^>]*/>)'
        r'(<w:t(?:\s[^>]*)?>)[^<]*(</w:t>)'
    )

    def repl(m):
        return m.group(1) + m.group(2) + val + m.group(3)
    return re.sub(pattern, repl, xml, flags=re.DOTALL)


def _renumber_ids(xml, offset):
    """Adauga offset la toate atributele w:id din XML."""
    return re.sub(
        r'w:id="(\d+)"',
        lambda m: f'w:id="{int(m.group(1)) + offset}"',
        xml
    )


def _unpack_docx(docx_path, dest_dir):
    """Despachetare .docx in directorul dest."""
    os.makedirs(dest_dir, exist_ok=True)
    with zipfile.ZipFile(docx_path, 'r') as z:
        z.extractall(dest_dir)


def _pack_docx(source_dir, output_path):
    """Impachetare directorul in .docx."""
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(source_dir):
            for f in files:
                full = os.path.join(root, f)
                arcname = os.path.relpath(full, source_dir)
                zf.write(full, arcname)


def preview_merge(master_path):
    """Previzualizeaza datele pentru mail merge."""
    wb = load_workbook(master_path, read_only=True)
    ws = wb.active
    total = 0
    uats = {}
    missing_fields = set()

    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        total += 1
        uat = str(ws.cell(r, 4).value or '').strip()
        uats[uat] = uats.get(uat, 0) + 1

        # Verifica campuri importante
        for col, label in [(30, 'Text HG'), (32, 'Decizie expropriere'),
                           (33, 'Decizie numire comisie'), (34, 'Membru1')]:
            if not ws.cell(r, col).value:
                missing_fields.add(label)

    wb.close()
    return {
        'total': total,
        'uats': uats,
        'missing_fields': list(missing_fields),
    }


def generate_word_merge(master_path, template_path, output_dir):
    """Genereaza document Word prin mail merge.
    Returneaza calea fisierului generat."""
    # 1. Citire MASTER
    wb = load_workbook(master_path)
    ws = wb.active

    rows_data = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        rows_data.append({f: _get_val(ws, r, c) for f, c in FIELD_COL.items()})

    if not rows_data:
        raise ValueError('MASTER-ul nu contine date de procesat.')

    # 2. Despachetare template
    unpacked_dir = os.path.join(output_dir, '_unpacked_template')
    if os.path.exists(unpacked_dir):
        shutil.rmtree(unpacked_dir)
    _unpack_docx(template_path, unpacked_dir)

    # 3. Citire template XML
    doc_xml_path = os.path.join(unpacked_dir, 'word', 'document.xml')
    with open(doc_xml_path, 'r', encoding='utf-8') as f:
        template_xml = f.read()

    header_xml_path = os.path.join(unpacked_dir, 'word', 'header1.xml')
    has_header = os.path.exists(header_xml_path)
    header_xml = ''
    if has_header:
        with open(header_xml_path, 'r', encoding='utf-8') as f:
            header_xml = f.read()

    # 4. Extrage body (fara sectPr)
    body_match = re.search(r'<w:body>(.*)</w:body>', template_xml, re.DOTALL)
    if not body_match:
        raise ValueError('Nu s-a gasit <w:body> in document.xml')
    body_content = body_match.group(1)

    sect_match = re.search(r'(<w:sectPr[\s>].*?</w:sectPr>)\s*$', body_content, re.DOTALL)
    if sect_match:
        sect_pr = sect_match.group(1)
        body_without_sect = body_content[:sect_match.start()].strip()
    else:
        sect_pr = ''
        body_without_sect = body_content.strip()

    # 5. Calculeaza ID maxim
    max_id = max(
        (int(m) for m in re.findall(r'w:id="(\d+)"', body_without_sect)),
        default=0
    )

    PAGE_BREAK = '\n<w:p><w:r><w:br w:type="page"/></w:r></w:p>\n'

    # 6. Genereaza sectiunile
    sections = []
    for i, row in enumerate(rows_data):
        sec = body_without_sect
        for field, value in row.items():
            sec = _replace_field(sec, field, value)
        sec = _renumber_ids(sec, i * (max_id + 10))
        sections.append(sec)

    combined_body = PAGE_BREAK.join(sections)
    if sect_pr:
        combined_body += '\n' + sect_pr

    new_xml = template_xml[:body_match.start(1)] + combined_body + template_xml[body_match.end(1):]

    # 7. Scrie output
    out_word_dir = os.path.join(output_dir, '_output_word')
    if os.path.exists(out_word_dir):
        shutil.rmtree(out_word_dir)
    shutil.copytree(unpacked_dir, out_word_dir)

    with open(os.path.join(out_word_dir, 'word', 'document.xml'), 'w', encoding='utf-8') as f:
        f.write(new_xml)

    # 8. Header (daca exista Antet_ in header1.xml)
    if has_header and 'Antet_' in header_xml:
        antet_val = _get_val(ws, 2, 31)
        if antet_val:
            antet_lines = antet_val.split('\n')
            runs_xml = ''
            for i, line in enumerate(antet_lines):
                runs_xml += f'<w:r><w:t xml:space="preserve">{_xml_escape(line)}</w:t>'
                if i < len(antet_lines) - 1:
                    runs_xml += '<w:br/>'
                runs_xml += '</w:r>'
            antet_pattern = (
                r'<w:fldChar[^>]+fldCharType="begin"[^>]*/>'
                r'.*?MERGEFIELD\s+Antet_.*?'
                r'<w:fldChar[^>]+fldCharType="end"[^>]*/>'
            )
            new_header = re.sub(antet_pattern, runs_xml, header_xml, flags=re.DOTALL)
            with open(os.path.join(out_word_dir, 'word', 'header1.xml'), 'w', encoding='utf-8') as f:
                f.write(new_header)

    # 9. Pack in DOCX
    output_docx = os.path.join(output_dir, 'H si PV - MASTER generat.docx')
    _pack_docx(out_word_dir, output_docx)

    # Cleanup temp dirs
    shutil.rmtree(unpacked_dir, ignore_errors=True)
    shutil.rmtree(out_word_dir, ignore_errors=True)

    return {
        'filename': os.path.basename(output_docx),
        'count': len(rows_data),
    }


def convert_doc_to_docx(doc_path, output_dir):
    """Converteste .doc in .docx. Returneaza calea .docx."""
    docx_name = os.path.splitext(os.path.basename(doc_path))[0] + '.docx'
    docx_path = os.path.join(output_dir, docx_name)

    try:
        import subprocess
        # Linux: LibreOffice
        result = subprocess.run(
            ['soffice', '--headless', '--convert-to', 'docx', '--outdir', output_dir, doc_path],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0 and os.path.exists(docx_path):
            return docx_path
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    try:
        import win32com.client
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        doc = word.Documents.Open(os.path.abspath(doc_path))
        doc.SaveAs(os.path.abspath(docx_path), FileFormat=16)
        doc.Close()
        word.Quit()
        return docx_path
    except ImportError:
        pass

    raise ValueError(
        'Nu se poate converti .doc in .docx. '
        'Incarcati direct un fisier .docx sau instalati LibreOffice / pywin32.'
    )
