import os
import re
import shutil
import zipfile
import pandas as pd
from openpyxl import load_workbook


def parse_borderou(xlsx_path):
    """Citeste un borderou (BR/BP) si extrage pozitiile HG si nr. hotarare.
    Returneaza dict cu 'filename', 'positions' (list of pozitie_hg int),
    'hotarari' (set of nr. hotarare str)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    positions = []
    hotarari = set()

    # Detect data start row (first row after header that has numeric value in col B)
    data_start = None
    for row in range(1, ws.max_row + 1):
        val_b = ws.cell(row=row, column=2).value
        if val_b is not None:
            try:
                int(float(val_b))
                if data_start is None:
                    data_start = row
            except (ValueError, TypeError):
                continue

    if data_start is None:
        wb.close()
        return {'filename': os.path.basename(xlsx_path), 'positions': [], 'hotarari': set()}

    for row in range(data_start, ws.max_row + 1):
        val_b = ws.cell(row=row, column=2).value  # Pozitie HG
        val_f = ws.cell(row=row, column=6).value  # Nr. Hotarare

        # Stop at TOTAL row
        for col in range(1, 13):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and 'TOTAL' in str(cell_val).upper():
                wb.close()
                return {
                    'filename': os.path.basename(xlsx_path),
                    'positions': positions,
                    'hotarari': hotarari,
                }

        if val_b is not None:
            try:
                poz = int(float(val_b))
                positions.append(poz)
            except (ValueError, TypeError):
                continue

            if val_f is not None:
                nr_hot = str(val_f).strip()
                if nr_hot:
                    hotarari.add(nr_hot)

    wb.close()
    return {
        'filename': os.path.basename(xlsx_path),
        'positions': positions,
        'hotarari': hotarari,
    }


def scan_hpv_files(file_paths):
    """Scaneaza fisierele Hot/PV si extrage nr. hotarare din nume.
    Returneaza dict {nr_hotarare_str: filepath}."""
    lookup = {}
    for fp in file_paths:
        basename = os.path.splitext(os.path.basename(fp))[0]
        # Pattern: "H si PV nr. 4", "H si PV nr 4", "Hot 4", "PV 4", "nr. 4", "nr 4"
        # Also match just a number at the end
        m = re.search(r'nr\.?\s*(\d+)', basename, re.IGNORECASE)
        if m:
            nr = m.group(1)
            lookup[nr] = fp
        else:
            # Try to find any number in the name
            m = re.search(r'(\d+)', basename)
            if m:
                nr = m.group(1)
                lookup[nr] = fp
    return lookup


def scan_doc_cadastrale(file_paths, folder_paths=None):
    """Scaneaza doc. cadastrale si extrage pozitia HG din nume.
    Accepta atat fisiere cat si foldere.
    Formate: 'p 1245.pdf', 'poz 1245.pdf', '1245.pdf', folder '1245_xxx'
    Returneaza dict {pozitie_hg_int: {'path': filepath_or_folder, 'is_folder': bool, 'name': original_name}}."""
    lookup = {}
    if folder_paths is None:
        folder_paths = []

    all_items = [(fp, False) for fp in file_paths] + [(fp, True) for fp in folder_paths]

    for item_path, is_folder in all_items:
        basename = os.path.basename(item_path)
        name_no_ext = os.path.splitext(basename)[0] if not is_folder else basename

        # Pattern: "p 1245", "poz 1245", "poz. 1245", "P 1245"
        m = re.match(r'^(?:poz\.?|p)\s*(\d+)', name_no_ext, re.IGNORECASE)
        if m:
            poz = int(m.group(1))
            lookup[poz] = {'path': item_path, 'is_folder': is_folder, 'name': basename}
            continue

        # Pattern: starts with number "1245" or "1245_xxx"
        m = re.match(r'^(\d+)', name_no_ext)
        if m:
            poz = int(m.group(1))
            lookup[poz] = {'path': item_path, 'is_folder': is_folder, 'name': basename}

    return lookup


def build_matching_preview(borderou_data_list, hpv_lookup, doc_cad_lookup):
    """Construieste preview-ul matching-ului pentru afisare.
    Returneaza lista de dicts per borderou cu info matching."""
    preview = []
    for bd in borderou_data_list:
        # Match Hot/PV
        matched_hpv = {}
        missing_hpv = set()
        for nr_hot in bd['hotarari']:
            if nr_hot in hpv_lookup:
                matched_hpv[nr_hot] = os.path.basename(hpv_lookup[nr_hot])
            else:
                missing_hpv.add(nr_hot)

        # Match Doc Cadastrale
        matched_doc = {}
        missing_doc = []
        for poz in bd['positions']:
            if poz in doc_cad_lookup:
                matched_doc[poz] = doc_cad_lookup[poz]['name']
            else:
                missing_doc.append(poz)

        preview.append({
            'filename': bd['filename'],
            'total_positions': len(bd['positions']),
            'total_hotarari': len(bd['hotarari']),
            'matched_hpv': matched_hpv,
            'missing_hpv': list(missing_hpv),
            'matched_doc': matched_doc,
            'missing_doc': missing_doc,
        })

    return preview


def organize_files(borderou_data_list, hpv_lookup, doc_cad_lookup, borderou_paths, output_dir):
    """Creeaza folderele si copiaza fisierele. Returneaza summary."""
    results = []

    for bd, br_path in zip(borderou_data_list, borderou_paths):
        folder_name = os.path.splitext(bd['filename'])[0]
        folder_path = os.path.join(output_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Copy the borderou itself
        shutil.copy2(br_path, os.path.join(folder_path, bd['filename']))

        # Copy Hot/PV files
        hpv_dir = os.path.join(folder_path, 'Hot si PV')
        os.makedirs(hpv_dir, exist_ok=True)
        copied_hpv = 0
        for nr_hot in bd['hotarari']:
            if nr_hot in hpv_lookup:
                src = hpv_lookup[nr_hot]
                dst = os.path.join(hpv_dir, os.path.basename(src))
                if not os.path.exists(dst):
                    shutil.copy2(src, dst)
                    copied_hpv += 1

        # Copy Doc Cadastrale
        doc_dir = os.path.join(folder_path, 'Doc cadastrale')
        os.makedirs(doc_dir, exist_ok=True)
        copied_doc = 0
        for poz in bd['positions']:
            if poz in doc_cad_lookup:
                info = doc_cad_lookup[poz]
                if info['is_folder']:
                    dst = os.path.join(doc_dir, info['name'])
                    if not os.path.exists(dst):
                        shutil.copytree(info['path'], dst)
                        copied_doc += 1
                else:
                    dst = os.path.join(doc_dir, os.path.basename(info['path']))
                    if not os.path.exists(dst):
                        shutil.copy2(info['path'], dst)
                        copied_doc += 1

        results.append({
            'folder_name': folder_name,
            'filename': bd['filename'],
            'total_positions': len(bd['positions']),
            'copied_hpv': copied_hpv,
            'copied_doc': copied_doc,
            'missing_hpv': len(bd['hotarari']) - copied_hpv,
            'missing_doc': len(bd['positions']) - sum(1 for p in bd['positions'] if p in doc_cad_lookup),
        })

    return results


def create_output_zip(output_dir, zip_name):
    """Creeaza un ZIP cu toate folderele din output_dir."""
    zip_path = os.path.join(output_dir, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(output_dir):
            for f in files:
                if f == zip_name:
                    continue
                full_path = os.path.join(root, f)
                arcname = os.path.relpath(full_path, output_dir)
                zf.write(full_path, arcname)
    return zip_path


def extract_zip_contents(zip_path, extract_dir):
    """Extrage un ZIP si returneaza lista de fisiere si foldere."""
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(extract_dir)

    files = []
    folders = []
    for item in os.listdir(extract_dir):
        full_path = os.path.join(extract_dir, item)
        if os.path.isdir(full_path):
            # Check if it contains subdirectories (nested zip structure)
            sub_items = os.listdir(full_path)
            has_subdirs = any(os.path.isdir(os.path.join(full_path, s)) for s in sub_items)
            if has_subdirs:
                # This might be a parent folder containing doc cad folders
                for sub in sub_items:
                    sub_path = os.path.join(full_path, sub)
                    if os.path.isdir(sub_path):
                        folders.append(sub_path)
                    else:
                        files.append(sub_path)
            else:
                folders.append(full_path)
        else:
            files.append(full_path)

    return files, folders
