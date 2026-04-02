import pandas as pd
from openpyxl import load_workbook
from copy import copy
import os


def analyze_master(master_path):
    """Citeste MASTER si returneaza date per sheet/UAT."""
    xls = pd.ExcelFile(master_path)
    result = {}
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        # Find DATA column (scan header row for 'DATA')
        date_col = 40  # default AO
        for j in range(df.shape[1]):
            if pd.notna(df.iloc[0, j]) and 'DATA' in str(df.iloc[0, j]).upper():
                date_col = j
                break
        entries = []
        for i in range(1, len(df)):
            poz = df.iloc[i, 1]  # col B = POZITIE HG
            nr_pv = df.iloc[i, 0]  # col A = Nr. PV si Hot
            data = df.iloc[i, date_col]
            if pd.notna(poz) and pd.notna(nr_pv):
                entries.append({
                    'pozitie_hg': int(poz),
                    'pv_nr': nr_pv,
                    'data': str(data) if pd.notna(data) else '',
                })
        result[sheet_name] = entries
    return result


def update_situatie_col_n(situatie_path, master_data, output_path):
    """Actualizeaza col N din SITUATIE cu PV/data din MASTER. Salveaza in output_path."""
    # Build lookup from all sheets
    lookup = {}
    for sheet_name, entries in master_data.items():
        for e in entries:
            lookup[e['pozitie_hg']] = f"{e['pv_nr']} / {e['data']}"

    wb = load_workbook(situatie_path)
    ws = wb.active
    updated = 0
    for row in range(2, ws.max_row + 1):
        poz_val = ws.cell(row=row, column=1).value
        if poz_val is not None:
            try:
                poz_int = int(poz_val)
            except (ValueError, TypeError):
                continue
            if poz_int in lookup:
                ws.cell(row=row, column=14).value = lookup[poz_int]
                updated += 1
    wb.save(output_path)
    return {'updated': updated, 'total_master': len(lookup)}


def analyze_situatie(situatie_path):
    """Analizeaza SITUATIE si returneaza statistici."""
    sit = pd.read_excel(situatie_path, header=None, sheet_name=0)
    total = len(sit) - 1  # minus header

    rows = []
    for i in range(1, len(sit)):
        n_val = sit.iloc[i, 13]
        t_val = sit.iloc[i, 19]
        has_n = pd.notna(n_val) and '/' in str(n_val)
        has_t = pd.notna(t_val) and str(t_val).strip() != ''
        if not (has_n and has_t):
            continue
        d_val = str(sit.iloc[i, 3]).strip().upper() if pd.notna(sit.iloc[i, 3]) else ''
        t_str = str(t_val).strip().upper()
        uat = str(sit.iloc[i, 2]).strip() if pd.notna(sit.iloc[i, 2]) else ''
        if not uat:
            continue
        rows.append({
            'poz': sit.iloc[i, 0],
            'uat': uat,
            'judet': str(sit.iloc[i, 1]).strip() if pd.notna(sit.iloc[i, 1]) else '',
            'match': d_val == t_str,
        })

    uats = sorted(set(r['uat'] for r in rows))
    uat_stats = []
    for uat in uats:
        uat_rows = [r for r in rows if r['uat'] == uat]
        uat_stats.append({
            'uat': uat,
            'judet': uat_rows[0]['judet'] if uat_rows else '',
            'total': len(uat_rows),
            'match': sum(1 for r in uat_rows if r['match']),
            'mismatch': sum(1 for r in uat_rows if not r['match']),
        })

    return {
        'total_situatie': total,
        'total_procesabile': len(rows),
        'total_match': sum(1 for r in rows if r['match']),
        'total_mismatch': sum(1 for r in rows if not r['match']),
        'uats': uat_stats,
    }


def parse_recipise(recipise_path):
    """Citeste fisierul RECIPISE si returneaza lookup dict {(pozitie_hg, nr_cadastral): nr_recipisa}.
    Scaneaza header-ul pentru a detecta automat coloanele."""
    df = pd.read_excel(recipise_path, header=None, sheet_name=0)
    if df.empty:
        return {}

    # Scan header row for column indices
    poz_col = None
    cad_col = None
    rec_col = None
    header_row = 0

    for i in range(min(5, len(df))):
        for j in range(df.shape[1]):
            val = str(df.iloc[i, j]).strip().upper() if pd.notna(df.iloc[i, j]) else ''
            if not val:
                continue
            if any(k in val for k in ['POZIT', 'POZ.', 'POZ ', 'NR. CRT', 'NR.CRT', 'POZITIE', 'POZITIA', 'POZIȚIA', 'POZIŢIA']):
                if 'HG' in val or poz_col is None:
                    poz_col = j
                    header_row = i
            if any(k in val for k in ['CADASTRAL', 'NR. CAD', 'NR.CAD', 'NR CAD', 'NUMAR CADASTRAL']):
                cad_col = j
                header_row = i
            if any(k in val for k in ['RECIPIS', 'NR. RECIPIS', 'NR.RECIPIS', 'NR RECIPIS', 'RECIPISA']):
                rec_col = j
                header_row = i

    if poz_col is None or cad_col is None or rec_col is None:
        raise ValueError(
            f'Nu am gasit coloanele necesare in RECIPISE. '
            f'Pozitie HG: {"gasit" if poz_col is not None else "LIPSA"}, '
            f'Nr. cadastral: {"gasit" if cad_col is not None else "LIPSA"}, '
            f'Nr. recipisa: {"gasit" if rec_col is not None else "LIPSA"}. '
            f'Verificati ca header-ul contine aceste coloane.'
        )

    lookup = {}
    for i in range(header_row + 1, len(df)):
        poz_val = df.iloc[i, poz_col]
        cad_val = df.iloc[i, cad_col]
        rec_val = df.iloc[i, rec_col]
        if pd.notna(poz_val) and pd.notna(cad_val) and pd.notna(rec_val):
            try:
                poz_key = int(float(poz_val))
            except (ValueError, TypeError):
                continue
            cad_key = str(cad_val).strip()
            lookup[(poz_key, cad_key)] = str(rec_val).strip()
    return lookup


def _parse_nr_data(s):
    if '/' in str(s):
        parts = str(s).split('/', 1)
        return parts[0].strip(), parts[1].strip()
    return str(s), ''


def _create_br(template_path, output_path, uat, judet, rows, br_num):
    """Genereaza un BR din template, pastrind formatarea."""
    wb = load_workbook(template_path)
    ws = wb.active

    # Determine max columns (10 base + 1 if recipise data exists)
    has_recipise = any(r.get('nr_recipisa') for r in rows)
    max_col = 11 if has_recipise else 10

    # Save styles from row 6
    row6_styles = {}
    for j in range(1, max_col + 1):
        cell = ws.cell(row=6, column=j)
        row6_styles[j] = {
            'font': copy(cell.font), 'fill': copy(cell.fill),
            'border': copy(cell.border), 'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
        }
    # For col K (11), fallback to col J style if not present in template
    if has_recipise and 11 not in row6_styles:
        row6_styles[11] = row6_styles[10]

    # Save total row styles
    total_styles = {}
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=8).value == 'TOTAL':
            for j in range(1, max_col + 1):
                cell = ws.cell(row=r, column=j)
                total_styles[j] = {
                    'font': copy(cell.font), 'fill': copy(cell.fill),
                    'border': copy(cell.border), 'alignment': copy(cell.alignment),
                    'number_format': cell.number_format,
                }
            break

    # Add header for recipise column if needed
    if has_recipise:
        header_cell = ws.cell(row=5, column=11, value='Nr. recipisa')
        if 10 in row6_styles:
            header_cell.font = copy(row6_styles[10]['font'])
            header_cell.border = copy(row6_styles[10]['border'])
            header_cell.alignment = copy(row6_styles[10]['alignment'])

    # Update title
    title_cell = ws.cell(row=3, column=1)
    if title_cell.value:
        title_cell.value = str(title_cell.value).replace('------', uat)

    # Clear data rows and unmerge
    for r in range(6, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            try:
                ws.cell(row=r, column=j).value = None
            except AttributeError:
                pass
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= 6:
            ws.unmerge_cells(str(m))

    # Write data
    ds = 6
    last_col_letter = 'K' if has_recipise else 'J'
    for idx, r in enumerate(rows):
        rn = ds + idx
        nr_hot, data_hot = _parse_nr_data(r['nr_data_hot'])
        vals = [
            idx + 1, r['poz'], r['suprafata'], r['nr_cadastral'],
            f"{r['uat']}, jud. {r['judet']}", nr_hot, data_hot,
            r['proprietar_d'] if br_num == '1' else r['proprietar_t'],
            r['valoare'], r['valoare']
        ]
        if has_recipise:
            vals.append(r.get('nr_recipisa', ''))
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rn, column=j, value=v)
            if j in row6_styles:
                cell.font = copy(row6_styles[j]['font'])
                cell.fill = copy(row6_styles[j]['fill'])
                cell.border = copy(row6_styles[j]['border'])
                cell.alignment = copy(row6_styles[j]['alignment'])
                cell.number_format = row6_styles[j]['number_format']

    # TOTAL row
    tr = ds + len(rows)
    for j in range(1, max_col + 1):
        cell = ws.cell(row=tr, column=j)
        if j in total_styles:
            cell.font = copy(total_styles[j]['font'])
            cell.fill = copy(total_styles[j]['fill'])
            cell.border = copy(total_styles[j]['border'])
            cell.alignment = copy(total_styles[j]['alignment'])
            cell.number_format = total_styles[j]['number_format']
    ws.cell(row=tr, column=8, value='TOTAL')
    ws.cell(row=tr, column=9, value=f'=SUM(I{ds}:I{tr-1})')
    ws.cell(row=tr, column=10, value=f'=SUM(J{ds}:J{tr-1})')

    # Footer
    fr = tr + 1
    ws.merge_cells(f'A{fr}:{last_col_letter}{fr}')
    ws.cell(row=fr, column=1, value='Not\u0103: prezentul borderou NU va fi urmat de plat\u0103.')
    ws.cell(row=fr, column=1).font = copy(row6_styles[1]['font'])
    ws.merge_cells(f'A{fr+2}:{last_col_letter}{fr+2}')
    ws.cell(row=fr+2, column=1, value='"PORTNOI \u0218I ASOCIA\u021aII" S.P.A.R.L. \nprin')
    ws.cell(row=fr+2, column=1).font = copy(row6_styles[1]['font'])
    ws.cell(row=fr+2, column=1).alignment = copy(row6_styles[1]['alignment'])
    ws.merge_cells(f'A{fr+3}:{last_col_letter}{fr+3}')
    ws.cell(row=fr+3, column=1, value='              Av. Ciprian-Gabriel Portnoi ')
    ws.cell(row=fr+3, column=1).font = copy(row6_styles[1]['font'])

    wb.save(output_path)
    return len(rows)


def generate_all_br(situatie_path, template_br1, template_br11, hg_number, output_dir, recipise_lookup=None):
    """Genereaza toate BR-urile. Returneaza lista de fisiere generate."""
    if recipise_lookup is None:
        recipise_lookup = {}
    os.makedirs(output_dir, exist_ok=True)
    sit = pd.read_excel(situatie_path, header=None, sheet_name=0)

    rows_data = []
    for i in range(1, len(sit)):
        n_val = sit.iloc[i, 13]
        t_val = sit.iloc[i, 19]
        if not (pd.notna(n_val) and '/' in str(n_val)):
            continue
        if not (pd.notna(t_val) and str(t_val).strip()):
            continue
        d_val = str(sit.iloc[i, 3]).strip().upper() if pd.notna(sit.iloc[i, 3]) else ''
        t_str = str(t_val).strip().upper()
        uat = str(sit.iloc[i, 2]).strip() if pd.notna(sit.iloc[i, 2]) else ''
        if not uat:
            continue
        poz_val = sit.iloc[i, 0]
        nr_cad = sit.iloc[i, 6] if pd.notna(sit.iloc[i, 6]) else ''
        nr_cad_str = str(nr_cad).strip()
        # Lookup recipisa by (pozitie_hg, nr_cadastral)
        try:
            poz_int = int(float(poz_val))
        except (ValueError, TypeError):
            poz_int = None
        nr_recipisa = ''
        if poz_int is not None and recipise_lookup:
            nr_recipisa = recipise_lookup.get((poz_int, nr_cad_str), '')
        rows_data.append({
            'poz': poz_val,
            'judet': str(sit.iloc[i, 1]).strip() if pd.notna(sit.iloc[i, 1]) else '',
            'uat': uat,
            'proprietar_d': str(sit.iloc[i, 3]).strip() if pd.notna(sit.iloc[i, 3]) else '',
            'proprietar_t': str(t_val).strip(),
            'nr_cadastral': nr_cad,
            'suprafata': sit.iloc[i, 11] if pd.notna(sit.iloc[i, 11]) else '',
            'valoare': sit.iloc[i, 12] if pd.notna(sit.iloc[i, 12]) else 0,
            'nr_data_hot': str(n_val).strip(),
            'match': d_val == t_str,
            'nr_recipisa': nr_recipisa,
        })

    matches = [r for r in rows_data if r['match']]
    mismatches = [r for r in rows_data if not r['match']]
    generated = []
    hg_clean = hg_number.replace('/', '-')

    # BR nr. 1 per UAT
    for uat in sorted(set(r['uat'] for r in matches)):
        ur = sorted([r for r in matches if r['uat'] == uat], key=lambda x: x['poz'])
        fn = f'BR nr. 1 UAT {uat} HG {hg_clean}.xlsx'
        n = _create_br(template_br1, os.path.join(output_dir, fn), uat, ur[0]['judet'], ur, '1')
        generated.append({'filename': fn, 'uat': uat, 'br_type': 'BR nr. 1', 'count': n})

    # BR nr. 1.1 per UAT
    for uat in sorted(set(r['uat'] for r in mismatches)):
        ur = sorted([r for r in mismatches if r['uat'] == uat], key=lambda x: x['poz'])
        fn = f'BR nr. 1.1 UAT {uat} HG {hg_clean}.xlsx'
        n = _create_br(template_br11, os.path.join(output_dir, fn), uat, ur[0]['judet'], ur, '1.1')
        generated.append({'filename': fn, 'uat': uat, 'br_type': 'BR nr. 1.1', 'count': n})

    return generated
